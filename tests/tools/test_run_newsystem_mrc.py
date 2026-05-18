"""Tests for tools/run_newsystem_mrc.py (Round 2 C3 — new-system runner stub)."""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

# ---------------------------------------------------------------------------
# Import helpers
# ---------------------------------------------------------------------------
TOOLS_DIR = Path(__file__).parent.parent.parent / "tools"
sys.path.insert(0, str(TOOLS_DIR))

from run_newsystem_mrc import main, MRC_SHEETS  # type: ignore[import]

REMIT_DATE = "2026-04-30"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _run(tmp_path: Path, mode: str = "pristine", seed: int = 42) -> Path:
    """Run the stub and return the out_dir Path."""
    out = tmp_path / mode
    rc = main([
        "--servicer", "mrc",
        "--remit-date", REMIT_DATE,
        "--out-dir", str(out),
        "--mode", mode,
        "--seed", str(seed),
    ])
    assert rc == 0, f"main() returned {rc} for mode={mode}"
    return out


# ---------------------------------------------------------------------------
# Test 1: Pristine mode → valid XLSX with 5 sheets matching canonical names
# ---------------------------------------------------------------------------

def test_pristine_five_sheets(tmp_path: Path) -> None:
    out = _run(tmp_path, "pristine")
    xlsx = out / "validation_report.xlsx"
    assert xlsx.exists(), "XLSX not produced"
    wb = openpyxl.load_workbook(str(xlsx))
    assert wb.sheetnames == MRC_SHEETS, (
        f"Sheet names mismatch.\nGot:      {wb.sheetnames}\nExpected: {MRC_SHEETS}"
    )


# ---------------------------------------------------------------------------
# Test 2: Pristine mode is deterministic (two runs → identical SHA256)
# ---------------------------------------------------------------------------

def test_pristine_deterministic(tmp_path: Path) -> None:
    out1 = _run(tmp_path / "run1", "pristine")
    out2 = _run(tmp_path / "run2", "pristine")

    meta1 = json.loads((out1 / "run_metadata.json").read_text())
    meta2 = json.loads((out2 / "run_metadata.json").read_text())

    sha1 = meta1["output"]["sha256"]
    sha2 = meta2["output"]["sha256"]
    assert sha1 == sha2, f"SHA256 mismatch between two pristine runs: {sha1!r} vs {sha2!r}"


# ---------------------------------------------------------------------------
# Test 3: Perturbed mode — perturbations.json has 4 entries and they are
#         reflected in the XLSX
# ---------------------------------------------------------------------------

def test_perturbed_four_perturbations(tmp_path: Path) -> None:
    out = _run(tmp_path, "perturbed")

    pert_path = out / "perturbations.json"
    assert pert_path.exists(), "perturbations.json not written in perturbed mode"

    perturbations = json.loads(pert_path.read_text())
    assert len(perturbations) == 4, (
        f"Expected 4 perturbations, got {len(perturbations)}: {perturbations}"
    )

    ids = {p["id"] for p in perturbations}
    assert ids == {"value_diff", "font_diff", "missing_row", "extra_sheet"}, (
        f"Unexpected perturbation ids: {ids}"
    )

    xlsx = out / "validation_report.xlsx"
    wb = openpyxl.load_workbook(str(xlsx))

    # 1. value_diff: MRC_General_Check row 2 'rate' column = 0.05
    ws_gen = wb["MRC_General_Check"]
    # find header row to locate 'rate' column
    headers = [ws_gen.cell(1, c).value for c in range(1, ws_gen.max_column + 1)]
    rate_col = headers.index("rate") + 1
    assert ws_gen.cell(2, rate_col).value == pytest.approx(0.05), (
        "value_diff not applied: rate in row 2 is not 0.05"
    )

    # 2. font_diff: MRC_Advance_Check A2 bold=True
    ws_adv = wb["MRC_Advance_Check"]
    assert ws_adv["A2"].font.bold is True, "font_diff not applied: A2 in Advance_Check not bold"

    # 3. missing_row: MRC_ServiceFee_Check has 3 data rows (not 4)
    ws_fee = wb["MRC_ServiceFee_Check"]
    n_data_rows = ws_fee.max_row - 1  # subtract header
    assert n_data_rows == 3, f"missing_row not applied: expected 3 data rows, got {n_data_rows}"

    # 4. extra_sheet: _PERTURBATION_EXTRA exists in workbook
    assert "_PERTURBATION_EXTRA" in wb.sheetnames, (
        "extra_sheet not applied: _PERTURBATION_EXTRA not found"
    )


# ---------------------------------------------------------------------------
# Test 4: Empty mode → valid minimal XLSX (sheets exist, no data rows)
# ---------------------------------------------------------------------------

def test_empty_mode_minimal_xlsx(tmp_path: Path) -> None:
    out = _run(tmp_path, "empty")
    xlsx = out / "validation_report.xlsx"
    assert xlsx.exists()
    wb = openpyxl.load_workbook(str(xlsx))
    # All 5 sheets must be present
    for name in MRC_SHEETS:
        assert name in wb.sheetnames, f"Sheet {name!r} missing in empty mode"
    # Each sheet has only a header row (no data rows)
    for name in MRC_SHEETS:
        ws = wb[name]
        assert ws.max_row == 1, (
            f"Sheet {name!r} should have 1 row (header only) in empty mode, "
            f"got {ws.max_row}"
        )


# ---------------------------------------------------------------------------
# Test 5: Metadata sidecar shape validation
# ---------------------------------------------------------------------------

def test_metadata_sidecar_shape(tmp_path: Path) -> None:
    out = _run(tmp_path, "pristine")
    meta_path = out / "run_metadata.json"
    assert meta_path.exists(), "run_metadata.json not produced"

    meta = json.loads(meta_path.read_text())

    # Required top-level keys
    required_keys = {
        "tool_version", "started_at", "finished_at", "system",
        "mode", "servicer", "remit_date", "output", "sheets",
    }
    missing = required_keys - meta.keys()
    assert not missing, f"run_metadata.json missing keys: {missing}"

    # output sub-object
    out_obj = meta["output"]
    assert "xlsx_path" in out_obj
    assert "sha256" in out_obj
    assert isinstance(out_obj["sha256"], str) and len(out_obj["sha256"]) == 64
    assert "size_bytes" in out_obj and isinstance(out_obj["size_bytes"], int)

    # sheets list
    assert isinstance(meta["sheets"], list) and len(meta["sheets"]) == 5
    for sheet_entry in meta["sheets"]:
        assert "name" in sheet_entry
        assert "rows" in sheet_entry
        assert "cols" in sheet_entry

    # system / mode
    assert meta["system"] == "newsystem-stub"
    assert meta["mode"] == "pristine"
    assert meta["servicer"] == "mrc"
    assert meta["remit_date"] == REMIT_DATE


# ---------------------------------------------------------------------------
# Test 6: --help exits 0 and mentions key flags
# ---------------------------------------------------------------------------

def test_help_exits_zero() -> None:
    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "run_newsystem_mrc.py"), "--help"],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, f"--help returned non-zero: {result.stderr}"
    output = result.stdout + result.stderr
    assert "--servicer" in output
    assert "--remit-date" in output
    assert "--mode" in output
    assert "--out-dir" in output
