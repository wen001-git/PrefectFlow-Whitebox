"""Tests for tools/xlsx_diff.py — cell-level XLSX diff tool."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color

# Make tools/ importable without installing as a package
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "tools"))

from xlsx_diff import (  # type: ignore[import]  # noqa: E402
    CAT_COL_WIDTH,
    CAT_FONT,
    CAT_FREEZE,
    CAT_MERGED,
    CAT_STRUCTURE,
    CAT_VALUE,
    MAJOR,
    MINOR,
    DiffOptions,
    diff_workbooks,
    exit_code_for,
    render_html,
)

# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


def _make_wb(cells: dict[str, object] | None = None) -> openpyxl.Workbook:
    """Create a simple single-sheet workbook with given cell values.

    cells: {"A1": value, "B2": value, ...}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if cells:
        for addr, val in cells.items():
            ws[addr] = val  # type: ignore[index]
    return wb


def _save(wb: openpyxl.Workbook, path: Path) -> None:
    wb.save(str(path))


# ---------------------------------------------------------------------------
# T1 — Identical workbooks → exit 0, no diffs
# ---------------------------------------------------------------------------


def test_identical_workbooks(tmp_path: Path) -> None:
    """Two identical workbooks must produce zero diffs and exit code 0."""
    wb = _make_wb({"A1": "hello", "B2": 42, "C3": 3.14})
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"
    _save(wb, leg)
    _save(wb, new)

    report = diff_workbooks(leg, new)
    assert report.identical
    assert report.major_count == 0
    assert report.minor_count == 0
    assert exit_code_for(report) == 0


# ---------------------------------------------------------------------------
# T2 — Single cell value diff → 1 major diff, exit 2
# ---------------------------------------------------------------------------


def test_single_value_diff(tmp_path: Path) -> None:
    """Changing one cell value → one major diff, exit code 2."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = _make_wb({"A1": 100})
    wb_n = _make_wb({"A1": 999})
    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    assert not report.identical
    value_diffs = [d for d in report.diffs if d.category == CAT_VALUE]
    assert len(value_diffs) == 1
    assert value_diffs[0].severity == MAJOR
    assert value_diffs[0].col == "A"
    assert value_diffs[0].row == 1
    assert exit_code_for(report) == 2


# ---------------------------------------------------------------------------
# T3 — Font change only → 1 minor diff, exit 1
# ---------------------------------------------------------------------------


def test_font_change_only(tmp_path: Path) -> None:
    """A bold-font change with same value → minor diff, exit 1."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    ws_l = wb_l.active
    ws_l["A1"] = "Hello"  # type: ignore[index]

    wb_n = openpyxl.Workbook()
    ws_n = wb_n.active
    ws_n["A1"] = "Hello"  # type: ignore[index]
    ws_n["A1"].font = Font(bold=True)  # type: ignore[index]

    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    assert not report.identical
    assert report.major_count == 0
    assert report.minor_count >= 1
    font_diffs = [d for d in report.diffs if d.category == CAT_FONT]
    assert any(d.note == "bold" for d in font_diffs)
    assert exit_code_for(report) == 1


# ---------------------------------------------------------------------------
# T4 — Missing sheet → reported, exit 2
# ---------------------------------------------------------------------------


def test_missing_sheet(tmp_path: Path) -> None:
    """Sheet present in legacy but absent from new → major structural diff."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    wb_l.active.title = "Sheet1"  # type: ignore[union-attr]
    wb_l.create_sheet("ExtraSheet")

    wb_n = openpyxl.Workbook()
    wb_n.active.title = "Sheet1"  # type: ignore[union-attr]

    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    struct_diffs = [d for d in report.diffs if d.category == CAT_STRUCTURE]
    assert any(d.sheet == "ExtraSheet" for d in struct_diffs)
    assert exit_code_for(report) == 2


# ---------------------------------------------------------------------------
# T5 — Extra sheet → reported, exit 2
# ---------------------------------------------------------------------------


def test_extra_sheet(tmp_path: Path) -> None:
    """Sheet absent in legacy but present in new → major structural diff."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    wb_l.active.title = "Sheet1"  # type: ignore[union-attr]

    wb_n = openpyxl.Workbook()
    wb_n.active.title = "Sheet1"  # type: ignore[union-attr]
    wb_n.create_sheet("BrandNewSheet")

    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    struct_diffs = [d for d in report.diffs if d.category == CAT_STRUCTURE]
    assert any(d.sheet == "BrandNewSheet" for d in struct_diffs)
    assert exit_code_for(report) == 2


# ---------------------------------------------------------------------------
# T6 — Merged-cells diff → reported
# ---------------------------------------------------------------------------


def test_merged_cells_diff(tmp_path: Path) -> None:
    """Difference in merged cell ranges → major diff."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    ws_l = wb_l.active
    ws_l.merge_cells("A1:C1")  # type: ignore[union-attr]
    ws_l["A1"] = "merged"  # type: ignore[index]

    wb_n = openpyxl.Workbook()
    ws_n = wb_n.active
    # no merge in new
    ws_n["A1"] = "merged"  # type: ignore[index]

    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    merged_diffs = [d for d in report.diffs if d.category == CAT_MERGED]
    assert len(merged_diffs) >= 1
    assert merged_diffs[0].severity == MAJOR
    assert exit_code_for(report) == 2


# ---------------------------------------------------------------------------
# T7 — Freeze-panes diff → reported
# ---------------------------------------------------------------------------


def test_freeze_panes_diff(tmp_path: Path) -> None:
    """Different freeze_panes setting → minor diff."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    ws_l = wb_l.active
    ws_l.freeze_panes = "A2"  # type: ignore[union-attr]
    ws_l["A1"] = "header"  # type: ignore[index]

    wb_n = openpyxl.Workbook()
    ws_n = wb_n.active
    ws_n.freeze_panes = None  # type: ignore[union-attr]
    ws_n["A1"] = "header"  # type: ignore[index]

    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    freeze_diffs = [d for d in report.diffs if d.category == CAT_FREEZE]
    assert len(freeze_diffs) == 1
    assert freeze_diffs[0].severity == MINOR
    assert freeze_diffs[0].legacy == "A2"


# ---------------------------------------------------------------------------
# T8 — Column-width diff → reported (minor)
# ---------------------------------------------------------------------------


def test_column_width_diff(tmp_path: Path) -> None:
    """Different column width → minor diff."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    ws_l = wb_l.active
    ws_l.column_dimensions["A"].width = 20.0  # type: ignore[union-attr]
    ws_l["A1"] = "data"  # type: ignore[index]

    wb_n = openpyxl.Workbook()
    ws_n = wb_n.active
    ws_n.column_dimensions["A"].width = 10.0  # type: ignore[union-attr]
    ws_n["A1"] = "data"  # type: ignore[index]

    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    col_diffs = [d for d in report.diffs if d.category == CAT_COL_WIDTH]
    assert len(col_diffs) == 1
    assert col_diffs[0].severity == MINOR
    assert col_diffs[0].col == "A"


# ---------------------------------------------------------------------------
# T9 — Float tolerance: 1.0 vs 1.0001 with tol=0.001 → identical
# ---------------------------------------------------------------------------


def test_float_tolerance_within_tol(tmp_path: Path) -> None:
    """Values within float tolerance → treated as equal, exit 0."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = _make_wb({"A1": 1.0})
    wb_n = _make_wb({"A1": 1.0001})
    _save(wb_l, leg)
    _save(wb_n, new)

    opts = DiffOptions(float_tolerance=0.001, ignore_style=True, ignore_format=True)
    report = diff_workbooks(leg, new, opts)
    value_diffs = [d for d in report.diffs if d.category == CAT_VALUE]
    assert len(value_diffs) == 0


def test_float_tolerance_exact_zero(tmp_path: Path) -> None:
    """Default tol=0.0: 1.0 vs 1.0001 → value diff."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = _make_wb({"A1": 1.0})
    wb_n = _make_wb({"A1": 1.0001})
    _save(wb_l, leg)
    _save(wb_n, new)

    opts = DiffOptions(float_tolerance=0.0, ignore_style=True, ignore_format=True)
    report = diff_workbooks(leg, new, opts)
    value_diffs = [d for d in report.diffs if d.category == CAT_VALUE]
    assert len(value_diffs) == 1


# ---------------------------------------------------------------------------
# T10 — --ignore-style suppresses font/fill/border diffs
# ---------------------------------------------------------------------------


def test_ignore_style_suppresses_font_fill(tmp_path: Path) -> None:
    """With ignore_style=True, font/fill/border diffs are NOT reported."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    ws_l = wb_l.active
    ws_l["A1"] = "Hello"  # type: ignore[index]

    wb_n = openpyxl.Workbook()
    ws_n = wb_n.active
    ws_n["A1"] = "Hello"  # type: ignore[index]
    ws_n["A1"].font = Font(bold=True, size=16, color="FF0000FF")  # type: ignore[index]
    ws_n["A1"].fill = PatternFill(  # type: ignore[index]
        fill_type="solid", fgColor=Color(rgb="FFFFFF00")
    )

    _save(wb_l, leg)
    _save(wb_n, new)

    opts = DiffOptions(ignore_style=True, ignore_format=True)
    report = diff_workbooks(leg, new, opts)
    style_diffs = [d for d in report.diffs if d.category in (CAT_FONT, "fill", "border")]
    assert len(style_diffs) == 0
    assert report.identical


# ---------------------------------------------------------------------------
# T11 — None vs 0 → value diff (not treated as equal)
# ---------------------------------------------------------------------------


def test_none_vs_zero_is_diff(tmp_path: Path) -> None:
    """None cell value vs 0 must be reported as a major diff."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = _make_wb({"A1": None})
    wb_n = _make_wb({"A1": 0})
    _save(wb_l, leg)
    _save(wb_n, new)

    opts = DiffOptions(ignore_style=True, ignore_format=True)
    report = diff_workbooks(leg, new, opts)
    value_diffs = [d for d in report.diffs if d.category == CAT_VALUE]
    assert len(value_diffs) == 1
    assert value_diffs[0].severity == MAJOR


# ---------------------------------------------------------------------------
# T12 — Sheet order difference is reported in structure
# ---------------------------------------------------------------------------


def test_sheet_order_diff(tmp_path: Path) -> None:
    """Different sheet order → structural diff in (workbook)."""
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = openpyxl.Workbook()
    wb_l.active.title = "Alpha"  # type: ignore[union-attr]
    wb_l.create_sheet("Beta")

    wb_n = openpyxl.Workbook()
    wb_n.active.title = "Beta"  # type: ignore[union-attr]
    wb_n.create_sheet("Alpha")

    _save(wb_l, leg)
    _save(wb_n, new)

    report = diff_workbooks(leg, new)
    workbook_diffs = [d for d in report.diffs if d.sheet == "(workbook)"]
    assert len(workbook_diffs) >= 1


# ---------------------------------------------------------------------------
# T13 — DiffReport serialisation round-trip
# ---------------------------------------------------------------------------


def test_report_to_dict_structure(tmp_path: Path) -> None:
    """to_dict() output contains all required top-level keys and correct types."""
    wb_l = _make_wb({"A1": 1, "B1": "foo"})
    wb_n = _make_wb({"A1": 2, "B1": "foo"})
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"
    _save(wb_l, leg)
    _save(wb_n, new)

    opts = DiffOptions(ignore_style=True, ignore_format=True)
    report = diff_workbooks(leg, new, opts)
    d = report.to_dict()

    assert d["tool_version"] == "1.0.0"
    assert "generated_at" in d
    assert "sha256" in d["legacy"]
    assert "sha256" in d["new"]
    assert isinstance(d["summary"]["identical"], bool)
    assert isinstance(d["summary"]["major_diff_count"], int)
    assert isinstance(d["summary"]["minor_diff_count"], int)
    assert isinstance(d["summary"]["per_sheet"], list)
    assert isinstance(d["diffs"], list)
    assert isinstance(d["structure"]["sheets_legacy"], list)
    assert isinstance(d["structure"]["sheets_new"], list)


# ---------------------------------------------------------------------------
# T14 — HTML render contains key fragments
# ---------------------------------------------------------------------------


def test_html_render_contains_fragments(tmp_path: Path) -> None:
    """render_html() produces valid HTML with expected fragments."""
    wb_l = _make_wb({"A1": 10})
    wb_n = _make_wb({"A1": 20})
    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"
    _save(wb_l, leg)
    _save(wb_n, new)

    opts = DiffOptions(ignore_style=True, ignore_format=True)
    report = diff_workbooks(leg, new, opts)
    html = render_html(report)

    assert "<!DOCTYPE html>" in html
    assert "xlsx_diff" in html
    assert "MAJOR" in html
    assert "badge-major" in html
    assert report.legacy_meta["sha256"][:8] in html


# ---------------------------------------------------------------------------
# T15 — NaN == NaN (no spurious diff)
# ---------------------------------------------------------------------------


def test_nan_equal_to_nan(tmp_path: Path) -> None:
    """NaN in both files → no value diff (NaN == NaN for diff purposes)."""

    leg = tmp_path / "leg.xlsx"
    new = tmp_path / "new.xlsx"

    wb_l = _make_wb({"A1": float("nan")})
    wb_n = _make_wb({"A1": float("nan")})
    _save(wb_l, leg)
    _save(wb_n, new)

    opts = DiffOptions(ignore_style=True, ignore_format=True)
    report = diff_workbooks(leg, new, opts)
    value_diffs = [d for d in report.diffs if d.category == CAT_VALUE]
    assert len(value_diffs) == 0
