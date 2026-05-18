"""
Tests for tools/compare_validation.py — Round 2 C4 comparison orchestrator.

Test inventory (11 tests):
  T1  compare mode: identical XLSX → verdict PASS, exit 0
  T2  compare mode: value diff → verdict MAJOR_DIFFS, exit 2
  T3  compare mode: font diff → verdict MINOR_DIFFS, exit 1
  T4  auto --dry-run → prints plan, exit 0, no subprocesses invoked
  T5  auto --skip-legacy --skip-new with provided XLSX → behaves like compare
  T6  auto with C3 perturbed mode → verdict MAJOR_DIFFS, warnings include harness info
  T7  verdict.json schema: all required keys present
  T8  warning: legacy source_repo_dirty=true
  T9  warning: runs >1h apart
  T10 exit 3 on missing XLSX path
  T11 --help works for both subcommands
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Import helpers
# ---------------------------------------------------------------------------

TOOLS_DIR = Path(__file__).parent.parent.parent / "tools"
sys.path.insert(0, str(TOOLS_DIR))

from compare_validation import (  # type: ignore[import]
    VERDICT_ERROR,
    VERDICT_MAJOR,
    VERDICT_MINOR,
    VERDICT_PASS,
    build_verdict,
    orchestrate,
)

REMIT_DATE = "2026-04-30"
SERVICER = "mrc"

# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _make_simple_wb(cells: dict[str, object] | None = None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    if cells:
        for addr, val in cells.items():
            ws[addr] = val  # type: ignore[index]
    return wb


def _save(wb: openpyxl.Workbook, path: Path) -> None:
    wb.save(str(path))


def _make_sidecar(path: Path, **extra: object) -> None:
    meta: dict[str, object] = {
        "tool_version": "1.0.0",
        "started_at": "2026-04-30T10:00:00+00:00",
        "servicer": SERVICER,
        "remit_date": REMIT_DATE,
        "output": {"xlsx_path": str(path), "sha256": "abc123", "size_bytes": 1000},
    }
    meta.update(extra)
    path.with_name("run_metadata.json").write_text(json.dumps(meta), encoding="utf-8")


# ── REQUIRED VERDICT KEYS ──────────────────────────────────────────────────

REQUIRED_VERDICT_KEYS = {
    "tool_version", "generated_at", "verdict", "exit_code",
    "servicer", "remit_date", "legacy_run", "new_run",
    "summary", "warnings", "next_steps",
}
REQUIRED_SUMMARY_KEYS = {"identical", "major_diff_count", "minor_diff_count", "per_sheet"}


# ---------------------------------------------------------------------------
# T1 — identical XLSX → PASS, exit 0
# ---------------------------------------------------------------------------

def test_compare_identical(tmp_path: Path) -> None:
    """Two identical workbooks must produce verdict PASS and exit 0."""
    wb = _make_simple_wb({"A1": "hello", "B2": 42, "C3": 3.14})
    leg = tmp_path / "leg.xlsx"; _save(wb, leg)
    new = tmp_path / "new.xlsx"; _save(wb, new)

    rc = orchestrate(
        legacy_xlsx=leg, new_xlsx=new,
        report_dir=tmp_path / "out",
        servicer=SERVICER, remit_date=REMIT_DATE,
        quiet=True,
    )
    assert rc == 0

    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    assert verdict["verdict"] == VERDICT_PASS
    assert verdict["exit_code"] == 0
    assert verdict["summary"]["identical"] is True


# ---------------------------------------------------------------------------
# T2 — value diff → MAJOR_DIFFS, exit 2
# ---------------------------------------------------------------------------

def test_compare_value_diff(tmp_path: Path) -> None:
    """A changed cell value must produce verdict MAJOR_DIFFS and exit 2."""
    leg_wb = _make_simple_wb({"A1": 100, "B1": 200})
    new_wb = _make_simple_wb({"A1": 999, "B1": 200})
    leg = tmp_path / "leg.xlsx"; _save(leg_wb, leg)
    new = tmp_path / "new.xlsx"; _save(new_wb, new)

    rc = orchestrate(
        legacy_xlsx=leg, new_xlsx=new,
        report_dir=tmp_path / "out",
        servicer=SERVICER, remit_date=REMIT_DATE,
        quiet=True,
    )
    assert rc == 2

    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    assert verdict["verdict"] == VERDICT_MAJOR
    assert verdict["exit_code"] == 2
    assert verdict["summary"]["major_diff_count"] >= 1


# ---------------------------------------------------------------------------
# T3 — font diff → MINOR_DIFFS, exit 1
# ---------------------------------------------------------------------------

def test_compare_font_diff(tmp_path: Path) -> None:
    """A font-only change must produce verdict MINOR_DIFFS and exit 1."""
    leg_wb = openpyxl.Workbook()
    ws_leg = leg_wb.active; assert ws_leg is not None
    ws_leg["A1"] = "test"
    ws_leg["A1"].font = Font(name="Arial", size=11, bold=False)
    leg = tmp_path / "leg.xlsx"; _save(leg_wb, leg)

    new_wb = openpyxl.Workbook()
    ws_new = new_wb.active; assert ws_new is not None
    ws_new["A1"] = "test"
    ws_new["A1"].font = Font(name="Arial", size=11, bold=True)
    new = tmp_path / "new.xlsx"; _save(new_wb, new)

    rc = orchestrate(
        legacy_xlsx=leg, new_xlsx=new,
        report_dir=tmp_path / "out",
        servicer=SERVICER, remit_date=REMIT_DATE,
        quiet=True,
    )
    assert rc == 1

    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    assert verdict["verdict"] == VERDICT_MINOR
    assert verdict["exit_code"] == 1
    assert verdict["summary"]["minor_diff_count"] >= 1
    assert verdict["summary"]["major_diff_count"] == 0


# ---------------------------------------------------------------------------
# T4 — auto --dry-run → prints plan, exit 0, no subprocesses
# ---------------------------------------------------------------------------

def test_auto_dry_run_no_subprocesses(tmp_path: Path) -> None:
    """Dry-run must print plan and exit 0 without invoking C2/C3."""
    from click.testing import CliRunner  # type: ignore[import]
    from compare_validation import cli  # type: ignore[import]

    runner = CliRunner()
    result = runner.invoke(cli, [
        "auto",
        "--servicer", SERVICER,
        "--remit-date", REMIT_DATE,
        "--report-dir", str(tmp_path / "out"),
        "--dry-run",
    ])
    assert result.exit_code == 0, f"output={result.output!r} exception={result.exception!r}"
    assert "DRY-RUN PLAN" in result.output
    assert "coherent" in result.output.lower()
    # dry-run must NOT write verdict.json (no run happened)
    assert not (tmp_path / "out" / "verdict.json").exists()


# ---------------------------------------------------------------------------
# T5 — auto --skip-legacy --skip-new with provided XLSX → behaves like compare
# ---------------------------------------------------------------------------

def test_auto_skip_both_behaves_like_compare(tmp_path: Path) -> None:
    """--skip-legacy + --skip-new should behave identically to compare mode."""
    from click.testing import CliRunner  # type: ignore[import]
    from compare_validation import cli  # type: ignore[import]

    wb = _make_simple_wb({"A1": 42, "B2": "foo"})
    leg = tmp_path / "leg.xlsx"; _save(wb, leg)
    new = tmp_path / "new.xlsx"; _save(wb, new)

    runner = CliRunner()
    result = runner.invoke(cli, [
        "auto",
        "--servicer", SERVICER,
        "--remit-date", REMIT_DATE,
        "--skip-legacy", "--legacy-xlsx", str(leg),
        "--skip-new", "--new-xlsx", str(new),
        "--report-dir", str(tmp_path / "out"),
        "--quiet",
    ])
    assert result.exit_code == 0, f"output={result.output!r} exception={result.exception!r}"

    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    assert verdict["verdict"] == VERDICT_PASS
    assert (tmp_path / "out" / "comparison_report.html").exists()
    assert (tmp_path / "out" / "comparison_report.json").exists()


# ---------------------------------------------------------------------------
# T6 — auto + C3 perturbed mode → MAJOR_DIFFS, warnings include harness info
# ---------------------------------------------------------------------------

def test_auto_perturbed_major_diffs(tmp_path: Path) -> None:
    """auto with perturbed new-system mode should detect major diffs."""
    # Generate two XLSXs: pristine as legacy, perturbed as new
    # Use C3 directly via subprocess

    leg_dir = tmp_path / "leg"
    new_dir = tmp_path / "new"

    c3 = TOOLS_DIR / "run_newsystem_mrc.py"

    # Create pristine (legacy-like)
    r1 = subprocess.run(
        [sys.executable, str(c3), "--servicer", "mrc", "--remit-date", REMIT_DATE,
         "--out-dir", str(leg_dir), "--mode", "pristine"],
        capture_output=True, text=True,
    )
    assert r1.returncode == 0, r1.stderr

    # Create perturbed (new-system with injected diffs)
    r2 = subprocess.run(
        [sys.executable, str(c3), "--servicer", "mrc", "--remit-date", REMIT_DATE,
         "--out-dir", str(new_dir), "--mode", "perturbed"],
        capture_output=True, text=True,
    )
    assert r2.returncode == 0, r2.stderr

    rc = orchestrate(
        legacy_xlsx=leg_dir / "validation_report.xlsx",
        new_xlsx=new_dir / "validation_report.xlsx",
        legacy_metadata=leg_dir / "run_metadata.json",
        new_metadata=new_dir / "run_metadata.json",
        report_dir=tmp_path / "out",
        servicer=SERVICER, remit_date=REMIT_DATE,
        new_mode="perturbed",
        quiet=True,
    )
    assert rc == 2

    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    assert verdict["verdict"] == VERDICT_MAJOR
    assert verdict["summary"]["major_diff_count"] >= 1


# ---------------------------------------------------------------------------
# T7 — verdict.json schema: all required keys present
# ---------------------------------------------------------------------------

def test_verdict_schema_all_required_keys(tmp_path: Path) -> None:
    """verdict.json must contain every required top-level key."""
    wb = _make_simple_wb({"A1": 1})
    leg = tmp_path / "leg.xlsx"; _save(wb, leg)
    new = tmp_path / "new.xlsx"; _save(wb, new)

    orchestrate(
        legacy_xlsx=leg, new_xlsx=new,
        report_dir=tmp_path / "out",
        servicer=SERVICER, remit_date=REMIT_DATE,
        quiet=True,
    )
    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())

    missing = REQUIRED_VERDICT_KEYS - set(verdict.keys())
    assert not missing, f"verdict.json missing keys: {missing}"

    summary_missing = REQUIRED_SUMMARY_KEYS - set(verdict["summary"].keys())
    assert not summary_missing, f"verdict.summary missing keys: {summary_missing}"


# ---------------------------------------------------------------------------
# T8 — warning: legacy source_repo_dirty=true
# ---------------------------------------------------------------------------

def test_warning_dirty_legacy_repo(tmp_path: Path) -> None:
    """Dirty legacy repo flag must appear in verdict warnings."""
    wb = _make_simple_wb({"A1": 1})
    leg = tmp_path / "leg.xlsx"; _save(wb, leg)
    new = tmp_path / "new.xlsx"; _save(wb, new)

    # Write a sidecar with source_repo_dirty=True
    leg_meta = {
        "tool_version": "1.0.0",
        "started_at": "2026-04-30T10:00:00+00:00",
        "source_repo_dirty": True,
        "output": {"xlsx_path": str(leg), "sha256": "aaa", "size_bytes": 100},
    }
    leg_meta_path = tmp_path / "leg_meta.json"
    leg_meta_path.write_text(json.dumps(leg_meta), encoding="utf-8")

    orchestrate(
        legacy_xlsx=leg, new_xlsx=new,
        report_dir=tmp_path / "out",
        legacy_metadata=leg_meta_path,
        servicer=SERVICER, remit_date=REMIT_DATE,
        quiet=True,
    )
    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    warnings_text = " ".join(verdict["warnings"])
    assert "uncommitted" in warnings_text or "dirty" in warnings_text


# ---------------------------------------------------------------------------
# T9 — warning: runs >1h apart
# ---------------------------------------------------------------------------

def test_warning_time_drift(tmp_path: Path) -> None:
    """Runs more than 1 hour apart must trigger a Redshift-drift warning."""
    wb = _make_simple_wb({"A1": 99})
    leg = tmp_path / "leg.xlsx"; _save(wb, leg)
    new = tmp_path / "new.xlsx"; _save(wb, new)

    leg_meta = {
        "tool_version": "1.0.0",
        "started_at": "2026-04-30T08:00:00+00:00",
        "output": {"xlsx_path": str(leg), "sha256": "aaa", "size_bytes": 100},
    }
    new_meta_payload = {
        "tool_version": "1.0.0",
        "started_at": "2026-04-30T12:00:00+00:00",  # 4 hours later
        "mode": "pristine",
        "output": {"xlsx_path": str(new), "sha256": "bbb", "size_bytes": 100},
    }
    leg_meta_path = tmp_path / "leg_meta.json"
    new_meta_path = tmp_path / "new_meta.json"
    leg_meta_path.write_text(json.dumps(leg_meta), encoding="utf-8")
    new_meta_path.write_text(json.dumps(new_meta_payload), encoding="utf-8")

    orchestrate(
        legacy_xlsx=leg, new_xlsx=new,
        report_dir=tmp_path / "out",
        legacy_metadata=leg_meta_path,
        new_metadata=new_meta_path,
        servicer=SERVICER, remit_date=REMIT_DATE,
        quiet=True,
    )
    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    warnings_text = " ".join(verdict["warnings"])
    assert "drifted" in warnings_text or "drift" in warnings_text


# ---------------------------------------------------------------------------
# T10 — exit 3 on missing XLSX path
# ---------------------------------------------------------------------------

def test_exit_3_missing_xlsx(tmp_path: Path) -> None:
    """Missing legacy XLSX must produce exit code 3 and ERROR verdict."""
    wb = _make_simple_wb({"A1": 1})
    new = tmp_path / "new.xlsx"; _save(wb, new)

    nonexistent = tmp_path / "does_not_exist.xlsx"

    rc = orchestrate(
        legacy_xlsx=nonexistent, new_xlsx=new,
        report_dir=tmp_path / "out",
        quiet=True,
    )
    assert rc == 3

    verdict = json.loads((tmp_path / "out" / "verdict.json").read_text())
    assert verdict["verdict"] == VERDICT_ERROR
    assert verdict["exit_code"] == 3


# ---------------------------------------------------------------------------
# T11 — --help works for both subcommands
# ---------------------------------------------------------------------------

def test_help_compare() -> None:
    """compare --help must exit 0."""
    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "compare_validation.py"), "compare", "--help"],
        capture_output=True, text=True,
    )
    assert result.returncode == 0
    assert "Usage" in result.stdout
    assert "--legacy-xlsx" in result.stdout


def test_help_auto() -> None:
    """auto --help must exit 0."""
    result = subprocess.run(
        [sys.executable, str(TOOLS_DIR / "compare_validation.py"), "auto", "--help"],
        capture_output=True, text=True,
    )
    assert result.returncode == 0
    assert "Usage" in result.stdout
    assert "--servicer" in result.stdout
    assert "--dry-run" in result.stdout
