"""Workbook-level renderer tests — structure, sheet names, freeze, widths."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from whitebox.renderer import render_workbook, write_workbook
from whitebox.renderer.formats import DEFAULT_COLUMN_WIDTH
from whitebox.sheets.base import SheetModel


def test_render_empty_returns_workbook_with_default_sheet() -> None:
    wb = render_workbook([])
    # Empty input: keep the openpyxl-default sheet so callers always
    # receive a valid Workbook (rendering nothing is not an error).
    assert len(wb.sheetnames) == 1


def test_render_creates_expected_sheets(minimal_models: list[SheetModel]) -> None:
    wb = render_workbook(minimal_models)
    assert wb.sheetnames == ["Alpha", "Beta"]


def test_header_and_body_row_counts(minimal_models: list[SheetModel]) -> None:
    wb = render_workbook(minimal_models)
    alpha = wb["Alpha"]
    # 1 header + 2 body rows = 3
    assert alpha.max_row == 3
    assert alpha.max_column == 4
    beta = wb["Beta"]
    assert beta.max_row == 2
    assert beta.max_column == 2


def test_freeze_panes_applied(minimal_models: list[SheetModel]) -> None:
    wb = render_workbook(minimal_models)
    assert wb["Alpha"].freeze_panes == "B2"
    assert wb["Beta"].freeze_panes == "A2"


def test_default_column_width_when_not_overridden(
    minimal_models: list[SheetModel],
) -> None:
    wb = render_workbook(minimal_models)
    alpha = wb["Alpha"]
    # column A explicitly set to 15 in ColumnSpec.width
    assert alpha.column_dimensions["A"].width == 15
    # column B has no width override → default
    assert alpha.column_dimensions["B"].width == DEFAULT_COLUMN_WIDTH


def test_write_workbook_round_trip(
    tmp_path: Path, minimal_models: list[SheetModel]
) -> None:
    wb = render_workbook(minimal_models)
    out = write_workbook(wb, tmp_path / "out.xlsx")
    assert out.exists()
    reloaded = openpyxl.load_workbook(out)
    assert reloaded.sheetnames == ["Alpha", "Beta"]
    assert reloaded["Alpha"].cell(row=1, column=1).value == "name"


def test_register_named_styles_is_idempotent(
    minimal_models: list[SheetModel],
) -> None:
    # Re-rendering must not raise (Workbook.add_named_style errors on dupes).
    render_workbook(minimal_models)
    render_workbook(minimal_models)


@pytest.mark.parametrize("name", ["header", "money", "money-int", "date", "str"])
def test_primary_named_styles_present(
    name: str, minimal_models: list[SheetModel]
) -> None:
    wb = render_workbook(minimal_models)
    assert name in {ns.name for ns in wb._named_styles}
