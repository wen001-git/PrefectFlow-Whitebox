"""Per-sheet structural invariants — must hold even when data is DEGRADED.

Even in the worst-case data path (every validator hits
``DEGRADED`` because its SQL could not execute), the engine must
still produce all five MRC sheets, in the documented order, with the
documented column headers. A regression here means the renderer or
sheet-builder layer broke its "always produce structurally-identical
shell" contract.
"""

from __future__ import annotations

from typing import Any

import openpyxl
import pytest

EXPECTED_SHEET_ORDER = (
    "MRC_Summary_check",
    "MRC_General_Check",
    "MRC_Advance_Check",
    "MRC_ServiceFee_Check",
    "MRC_Adv_Info",
)

# Header-row column counts derived from
# whitebox/sheets/mrc/_helpers.py (see docs/mrc/sheets.{zh,en}.md
# § 1.3 — per-sheet column-list helpers).
EXPECTED_MIN_COLUMNS: dict[str, int] = {
    "MRC_Summary_check": 14,
    "MRC_General_Check": 35,
    "MRC_Advance_Check": 27,
    "MRC_ServiceFee_Check": 8,
    "MRC_Adv_Info": 7,
}


@pytest.mark.acceptance
def test_all_five_sheets_present_in_order(engine_run: Any) -> None:
    wb = openpyxl.load_workbook(engine_run.xlsx_path, data_only=True)
    assert tuple(wb.sheetnames) == EXPECTED_SHEET_ORDER


@pytest.mark.acceptance
@pytest.mark.parametrize("sheet_name", list(EXPECTED_SHEET_ORDER))
def test_each_sheet_has_expected_header_width(
    engine_run: Any, sheet_name: str
) -> None:
    wb = openpyxl.load_workbook(engine_run.xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1] if c.value is not None]
    expected = EXPECTED_MIN_COLUMNS[sheet_name]
    assert len(headers) >= expected, (
        f"sheet {sheet_name} produced {len(headers)} headers; expected >= {expected}. "
        f"Headers: {headers}"
    )


@pytest.mark.acceptance
@pytest.mark.parametrize("sheet_name", list(EXPECTED_SHEET_ORDER))
def test_each_sheet_header_row_is_non_empty(
    engine_run: Any, sheet_name: str
) -> None:
    wb = openpyxl.load_workbook(engine_run.xlsx_path, data_only=True)
    ws = wb[sheet_name]
    first_cell = ws.cell(row=1, column=1).value
    assert first_cell not in (None, ""), (
        f"sheet {sheet_name} row 1 col 1 is empty — header missing"
    )
