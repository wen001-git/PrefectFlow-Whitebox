"""Fixture builders for integration comparison scenarios.

All builders are deterministic given a seed. Tests must call builders once and
reuse the output — builders never embed timestamps or random content inside cells.

Cross-ref: plan.md § 9.2 (G2b-LIVE C5 gate); test_harness_e2e.py.
"""
from __future__ import annotations

import random
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# Sheet definitions used by build_baseline_workbook
# ---------------------------------------------------------------------------

_SHEET_DEFS: list[tuple[str, list[str]]] = [
    ("Summary",    ["col_a", "col_b", "col_c", "col_d"]),
    ("General",    ["col_a", "col_b", "col_c", "col_d"]),
    ("Advance",    ["col_a", "col_b", "col_c", "col_d"]),
    ("ServiceFee", ["col_a", "col_b", "col_c", "col_d"]),
    ("AdvInfo",    ["col_a", "col_b", "col_c", "col_d"]),
]

# Number of data rows per sheet (excluding header)
_DATA_ROWS = 5


# ---------------------------------------------------------------------------
# Public builders
# ---------------------------------------------------------------------------


def build_baseline_workbook(path: Path, seed: int = 42) -> None:
    """Build a deterministic 5-sheet plausible workbook.

    Each sheet: 1 header row + 5 data rows × 4 numeric columns.
    Values are deterministic given *seed*.

    Args:
        path: Destination .xlsx file path (parent must exist).
        seed: RNG seed for determinism.
    """
    rng = random.Random(seed)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    for sheet_name, headers in _SHEET_DEFS:
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, hdr in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=hdr)
        for row in range(2, 2 + _DATA_ROWS):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col, value=round(rng.uniform(1.0, 100.0), 4))

    wb.save(str(path))


def build_value_diff(path: Path, baseline_path: Path) -> None:
    """Copy baseline; change Summary sheet B3 to a distinctive sentinel value.

    Expected diff: category=value, severity=major.
    """
    wb = openpyxl.load_workbook(str(baseline_path))
    ws = wb["Summary"]
    ws.cell(row=3, column=2).value = 9999.99  # sentinel — never in baseline
    wb.save(str(path))


def build_font_diff(path: Path, baseline_path: Path) -> None:
    """Copy baseline; bold Summary sheet A2 cell.

    Expected diff: category=font, severity=minor, note contains 'bold'.
    No value change — should produce MINOR_DIFFS exit 1.
    """
    wb = openpyxl.load_workbook(str(baseline_path))
    ws = wb["Summary"]
    ws.cell(row=2, column=1).font = Font(bold=True)
    wb.save(str(path))


def build_missing_sheet(path: Path, baseline_path: Path) -> None:
    """Copy baseline minus the AdvInfo sheet.

    Expected diff: category=structure, severity=major (sheet exists in legacy only).
    """
    wb = openpyxl.load_workbook(str(baseline_path))
    del wb["AdvInfo"]
    wb.save(str(path))


def build_extra_sheet(path: Path, baseline_path: Path) -> None:
    """Copy baseline plus one extra sheet _EXTRA_SHEET.

    Expected diff: category=structure, severity=major (sheet exists in new only).
    """
    wb = openpyxl.load_workbook(str(baseline_path))
    extra_ws = wb.create_sheet(title="_EXTRA_SHEET")
    extra_ws.cell(row=1, column=1, value="key")
    extra_ws.cell(row=1, column=2, value="value")
    extra_ws.cell(row=2, column=1, value="injected")
    extra_ws.cell(row=2, column=2, value=True)
    wb.save(str(path))


def build_merged_cells_diff(path: Path, baseline_path: Path) -> None:
    """Copy baseline; add merged range A2:B2 on Summary sheet.

    Expected diff: category=merged_cells, severity=major.
    """
    wb = openpyxl.load_workbook(str(baseline_path))
    ws = wb["Summary"]
    ws.merge_cells("A2:B2")
    wb.save(str(path))


def build_dimension_diff(path: Path, baseline_path: Path) -> None:
    """Copy baseline; add a 5th column to Summary sheet (col E).

    Summary baseline has 4 cols; the copy gets a 5th column with header + data.
    Expected diff: category=structure (sheet dimensions differ), severity=major.
    """
    rng = random.Random(999)  # fixed secondary seed for the extra column values
    wb = openpyxl.load_workbook(str(baseline_path))
    ws = wb["Summary"]
    ws.cell(row=1, column=5, value="col_e")  # extra header
    for row in range(2, 2 + _DATA_ROWS):
        ws.cell(row=row, column=5, value=round(rng.uniform(1.0, 100.0), 4))
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Reproducibility self-check (used by tests)
# ---------------------------------------------------------------------------


def workbook_value_fingerprint(path: Path) -> dict[str, Any]:
    """Return a dict of {sheet: {cell_addr: value}} for determinism assertions."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    result: dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells: dict[str, Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = cell.value
        result[sheet_name] = cells
    return result
