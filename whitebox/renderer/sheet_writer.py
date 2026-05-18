"""Per-sheet writer — owns the SheetModel → openpyxl Worksheet mapping.

Responsibilities (mirror legacy ``sheet_format``/``all_cell_format``/
``data_type_format``/``header_format``/``diff_cell_format``/
``column_auto_width`` — see ``util/gen_remit_validation_report.py``
lines 1694-1819, cited via ``docs/mrc/1.3-sheets.en.md`` § 3):

1. Header row with semantic ``header`` / ``header-diff`` styling.
2. Body rows with one primary style per cell + optional ``diff`` overlay.
3. Freeze panes (e.g. ``"A2"``) from :class:`SheetModel.freeze_panes`.
4. Merged cell ranges from :class:`SheetModel.merged_cells`.
5. Column widths (explicit ``ColumnSpec.width`` / ``column_widths`` map,
   else the chapter default of 20).
6. Sheet view defaults (gridlines on, zoom 100) — explicit so the
   resulting XLSX is deterministic and ``xlsx_diff`` does not flag a
   minor sheet_view delta.
"""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from whitebox.renderer.formats import DEFAULT_COLUMN_WIDTH
from whitebox.renderer.styles import OVERLAY_CLASSES, PRIMARY_CLASSES, STYLES
from whitebox.sheets.base import Cell, Row, SheetModel


def write_sheet(ws: Worksheet, model: SheetModel) -> None:
    """Populate ``ws`` from ``model``. Assumes named styles already registered."""
    _set_sheet_view(ws)
    _write_header(ws, model)
    _write_body(ws, model)
    _write_sections(ws, model)
    _apply_freeze(ws, model)
    _apply_merged_cells(ws, model)
    _apply_column_widths(ws, model)


# ---------------------------------------------------------------------------
# Sheet view — deterministic defaults so xlsx_diff doesn't flag drift.
# ---------------------------------------------------------------------------


def _set_sheet_view(ws: Worksheet) -> None:
    view = ws.sheet_view
    view.showGridLines = True
    view.zoomScale = 100
    view.zoomScaleNormal = 100


# ---------------------------------------------------------------------------
# Header + body row writers.
# ---------------------------------------------------------------------------


def _write_header(ws: Worksheet, model: SheetModel) -> None:
    if not model.columns:
        return
    for col_idx, col in enumerate(model.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.header_label)
        # ColumnSpec.highlight → header-diff; otherwise the model's
        # configured header_style (defaults to "header").
        if col.highlight:
            cell.style = "header-diff"
        else:
            primary = _select_primary(model.header_style.classes) or "header"
            cell.style = primary


def _write_body(ws: Worksheet, model: SheetModel) -> None:
    for row_offset, row in enumerate(model.rows, start=2):  # row 1 = header
        _write_row(ws, row, row_index=row_offset)


def _write_sections(ws: Worksheet, model: SheetModel) -> None:
    """Sections render after the flat body, separated by their title row.

    Per :class:`whitebox.sheets.base.SheetSection` docstring: sections
    are an optional grouping aid; MRC does not use them. The renderer
    writes the section title in column A (no styling beyond ``header``)
    and the section rows below it.
    """
    next_row = 2 + len(model.rows)
    for section in model.sections:
        if section.title:
            title_cell = ws.cell(row=next_row, column=1, value=section.title)
            title_cell.style = "header"
            next_row += 1
        for srow in section.rows:
            _write_row(ws, srow, row_index=next_row)
            next_row += 1


def _write_row(ws: Worksheet, row: Row, *, row_index: int) -> None:
    for col_idx, cell_data in enumerate(row.cells, start=1):
        target = ws.cell(row=row_index, column=col_idx, value=cell_data.value)
        _apply_cell_style(target, cell_data)


# ---------------------------------------------------------------------------
# Per-cell style cascade: primary class first, then overlays.
# ---------------------------------------------------------------------------


def _select_primary(classes: tuple[str, ...]) -> str | None:
    """Return the last primary class in ``classes``, or ``None``.

    "Last wins" matches legacy semantics — see ``coerce_money_cell``
    which appends ``money-int`` *after* the helper's default ``money``.
    """
    primary: str | None = None
    for cls in classes:
        if cls in PRIMARY_CLASSES:
            primary = cls
    return primary


def _select_overlays(classes: tuple[str, ...]) -> list[str]:
    return [cls for cls in classes if cls in OVERLAY_CLASSES]


def _apply_cell_style(target: object, cell: Cell) -> None:
    """Apply primary NamedStyle + overlay font/fill mutations to ``target``."""
    primary = _select_primary(cell.style.classes)
    if primary is not None:
        # Assigning ``cell.style`` by name requires the NamedStyle to be
        # registered on the workbook (done by render_workbook).
        target.style = primary  # type: ignore[attr-defined]

    # Honor the per-cell number_format override if set (e.g. helpers may
    # request a specific money format independent of the primary class).
    nf = cell.style.number_format
    if nf is not None:
        target.number_format = nf  # type: ignore[attr-defined]

    for overlay_name in _select_overlays(cell.style.classes):
        overlay = STYLES[overlay_name]
        if overlay.font is not None:
            # Construct a fresh Font instance so the named-style binding
            # is broken cleanly (openpyxl shares Font objects across
            # cells via NamedStyle; mutating in place would bleed).
            target.font = Font(  # type: ignore[attr-defined]
                name=overlay.font.name,
                size=overlay.font.size,
                color=overlay.font.color,
                bold=overlay.font.bold,
                italic=overlay.font.italic,
            )
        if overlay.fill is not None:
            target.fill = PatternFill(  # type: ignore[attr-defined]
                start_color=overlay.fill.start_color,
                end_color=overlay.fill.end_color,
                fill_type=overlay.fill.fill_type,
            )


# ---------------------------------------------------------------------------
# Freeze panes, merged cells, column widths.
# ---------------------------------------------------------------------------


def _apply_freeze(ws: Worksheet, model: SheetModel) -> None:
    if model.freeze_panes:
        ws.freeze_panes = model.freeze_panes


def _apply_merged_cells(ws: Worksheet, model: SheetModel) -> None:
    for range_str in model.merged_cells:
        ws.merge_cells(range_str)


def _apply_column_widths(ws: Worksheet, model: SheetModel) -> None:
    """Width precedence: ``column_widths`` map > ``ColumnSpec.width`` > default."""
    overrides = dict(model.column_widths)
    for col_idx, col in enumerate(model.columns, start=1):
        letter = get_column_letter(col_idx)
        width: int
        if col.key in overrides:
            width = int(overrides[col.key])
        elif col.width is not None:
            width = int(col.width)
        else:
            width = DEFAULT_COLUMN_WIDTH
        ws.column_dimensions[letter].width = float(width)


__all__ = ["write_sheet"]
