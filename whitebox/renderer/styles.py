"""Central semantic-class → openpyxl style registry.

This is the **only** place in the codebase that constructs ``Font``,
``PatternFill``, ``Alignment``, ``Border`` or ``NamedStyle`` instances
for the runtime renderer. Sheets emit semantic class names (see
:mod:`whitebox.sheets.base` module docstring for the frozen vocabulary)
and the renderer looks them up here.

Two flavors of class
--------------------

1. **Primary classes** (``header``, ``money``, ``money-int``, ``percent``,
   ``float``, ``int``, ``date``, ``str``, ``total``) — registered as
   :class:`openpyxl.styles.NamedStyle` on the workbook. The sheet writer
   selects exactly one of these per cell as the "base" style.

2. **Overlay classes** (``diff``, ``header-diff``, ``warning-red``,
   ``warning-grey``) — applied on top of a primary style by mutating
   ``cell.font`` / ``cell.fill`` after the primary is set. This mirrors
   the legacy ``diff_cell_format`` pass (1.3-sheets.en.md § 4.3) and
   avoids a combinatorial explosion of NamedStyle variants.

Note: ``header-diff`` is also registered as a NamedStyle so that
header cells can take a single ``cell.style = "header-diff"`` assignment
(legacy header_format does the same blue→pink swap as a single paint).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.workbook import Workbook

from whitebox.renderer.formats import (
    DEFAULT_FONT_COLOR,
    DEFAULT_FONT_NAME,
    DEFAULT_FONT_SIZE,
    DIFF_FILL_COLOR,
    DIFF_FONT_COLOR,
    DIFF_RELATION_FILL_COLOR,
    DIFF_RELATION_FONT_COLOR,
    GENERAL_FORMAT,
    HEADER_FILL_COLOR,
    MONEY_FORMAT,
    MONEY_INT_FORMAT,
    PERCENT_FORMAT,
)

# Thin black border on all four sides — applied to every cell to match
# the legacy default (1.3 § 4.2 "Default font" row).
_THIN_BLACK_SIDE = Side(border_style="thin", color="000000")
_DEFAULT_BORDER = Border(
    left=_THIN_BLACK_SIDE,
    right=_THIN_BLACK_SIDE,
    top=_THIN_BLACK_SIDE,
    bottom=_THIN_BLACK_SIDE,
)

_DEFAULT_ALIGNMENT = Alignment()  # openpyxl defaults — no overrides for body
_HEADER_ALIGNMENT = Alignment(horizontal="center")

_DEFAULT_FONT = Font(
    name=DEFAULT_FONT_NAME,
    size=DEFAULT_FONT_SIZE,
    color=DEFAULT_FONT_COLOR,
    bold=False,
    italic=False,
)
_HEADER_FONT = Font(
    name=DEFAULT_FONT_NAME,
    size=DEFAULT_FONT_SIZE,
    color=DEFAULT_FONT_COLOR,
    bold=True,
    italic=False,
)
_DIFF_HEADER_FONT = Font(
    name=DEFAULT_FONT_NAME,
    size=DEFAULT_FONT_SIZE,
    color=DIFF_FONT_COLOR,
    bold=True,
    italic=False,
)
_DIFF_BODY_FONT = Font(
    name=DEFAULT_FONT_NAME,
    size=DEFAULT_FONT_SIZE,
    color=DIFF_FONT_COLOR,
    bold=False,
    italic=False,
)
_WARNING_GREY_FONT = Font(
    name=DEFAULT_FONT_NAME,
    size=DEFAULT_FONT_SIZE,
    color=DIFF_RELATION_FONT_COLOR,
    bold=False,
    italic=False,
)

_NO_FILL = PatternFill(fill_type=None)
_HEADER_FILL = PatternFill(
    start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid"
)
_DIFF_FILL = PatternFill(
    start_color=DIFF_FILL_COLOR, end_color=DIFF_FILL_COLOR, fill_type="solid"
)
_DIFF_RELATION_FILL = PatternFill(
    start_color=DIFF_RELATION_FILL_COLOR,
    end_color=DIFF_RELATION_FILL_COLOR,
    fill_type="solid",
)


@dataclass(frozen=True)
class StyleDefinition:
    """Pure-data description of one semantic style class.

    ``role`` is either ``"primary"`` (rendered as a workbook-level
    :class:`NamedStyle`) or ``"overlay"`` (applied by mutating font/fill
    on top of an already-styled cell). ``font`` / ``fill`` may be
    ``None`` for overlays that only mutate one attribute.
    """

    name: str
    role: str  # "primary" | "overlay"
    font: Font | None
    fill: PatternFill | None
    alignment: Alignment
    border: Border
    number_format: str
    metadata: dict[str, str] = field(default_factory=dict)

    def to_named_style(self) -> NamedStyle:
        """Materialize this definition as an openpyxl ``NamedStyle``.

        Only valid for ``role == "primary"`` — overlays do not have a
        complete style and must be applied via the sheet writer.
        """
        if self.role != "primary":
            raise ValueError(
                f"StyleDefinition {self.name!r} has role={self.role!r}; "
                "only primary roles can be materialized as NamedStyle."
            )
        ns = NamedStyle(name=self.name)
        # font/fill are non-None for every primary class — these
        # assertions also keep mypy strict happy without a cast.
        assert self.font is not None
        assert self.fill is not None
        ns.font = self.font
        ns.fill = self.fill
        ns.alignment = self.alignment
        ns.border = self.border
        ns.number_format = self.number_format
        return ns


# ---------------------------------------------------------------------------
# Registry — frozen vocabulary, see whitebox.sheets.base module docstring.
# ---------------------------------------------------------------------------

STYLES: dict[str, StyleDefinition] = {
    "header": StyleDefinition(
        name="header",
        role="primary",
        font=_HEADER_FONT,
        fill=_HEADER_FILL,
        alignment=_HEADER_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    "header-diff": StyleDefinition(
        name="header-diff",
        role="primary",
        font=_DIFF_HEADER_FONT,
        fill=_DIFF_FILL,
        alignment=_HEADER_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    "money": StyleDefinition(
        name="money",
        role="primary",
        font=_DEFAULT_FONT,
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=MONEY_FORMAT,
    ),
    "money-int": StyleDefinition(
        name="money-int",
        role="primary",
        font=_DEFAULT_FONT,
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=MONEY_INT_FORMAT,
    ),
    "percent": StyleDefinition(
        name="percent",
        role="primary",
        font=_DEFAULT_FONT,
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=PERCENT_FORMAT,
    ),
    "float": StyleDefinition(
        name="float",
        role="primary",
        font=_DEFAULT_FONT,
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    "int": StyleDefinition(
        name="int",
        role="primary",
        font=_DEFAULT_FONT,
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    "date": StyleDefinition(
        name="date",
        role="primary",
        font=_DEFAULT_FONT,
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        # Legacy renderer does NOT set a number_format for date cells —
        # pandas/openpyxl serializes Python ``date`` to Excel's serial
        # and Excel falls back to its locale-default date display
        # (1.3 § 4.2 "all other types … values are written as-is").
        number_format=GENERAL_FORMAT,
    ),
    "str": StyleDefinition(
        name="str",
        role="primary",
        font=_DEFAULT_FONT,
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    "total": StyleDefinition(
        name="total",
        role="primary",
        font=_HEADER_FONT,  # bold totals row
        fill=_NO_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    # ---- overlays --------------------------------------------------------
    "diff": StyleDefinition(
        name="diff",
        role="overlay",
        font=_DIFF_BODY_FONT,
        fill=_DIFF_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    "warning-red": StyleDefinition(
        name="warning-red",
        role="overlay",
        font=_DIFF_BODY_FONT,
        fill=_DIFF_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
    "warning-grey": StyleDefinition(
        name="warning-grey",
        role="overlay",
        font=_WARNING_GREY_FONT,
        fill=_DIFF_RELATION_FILL,
        alignment=_DEFAULT_ALIGNMENT,
        border=_DEFAULT_BORDER,
        number_format=GENERAL_FORMAT,
    ),
}


# Primary classes in priority order: when a cell carries multiple
# semantic classes the *last* primary wins (matches the legacy "money
# wins over float wins over str" cascade — see helpers.coerce_money_cell
# behavior in whitebox/sheets/mrc/_helpers.py).
PRIMARY_CLASSES: tuple[str, ...] = tuple(
    name for name, spec in STYLES.items() if spec.role == "primary"
)
OVERLAY_CLASSES: tuple[str, ...] = tuple(
    name for name, spec in STYLES.items() if spec.role == "overlay"
)


def register_named_styles(wb: Workbook) -> None:
    """Add every primary :class:`StyleDefinition` to ``wb`` as a NamedStyle.

    Idempotent: if the workbook already has a NamedStyle with the same
    name (e.g. ``"Normal"`` shipped by openpyxl), the duplicate is
    silently skipped. This allows multiple :func:`render_workbook` calls
    against the same workbook without raising ``ValueError``.
    """
    existing = {ns.name for ns in wb._named_styles}
    for spec in STYLES.values():
        if spec.role != "primary":
            continue
        if spec.name in existing:
            continue
        wb.add_named_style(spec.to_named_style())


# ---------------------------------------------------------------------------
# Serialization — drives the golden styles_snapshot.json fixture.
# ---------------------------------------------------------------------------


def _serialize_color(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    # openpyxl ``Color`` exposes ``rgb`` (e.g. "00bccde9"), ``type``
    # ("rgb" | "theme" | "indexed") and possibly ``value``. We keep the
    # raw string when present; otherwise fall back to ``str(value)``.
    rgb = getattr(value, "rgb", None)
    if isinstance(rgb, str):
        return rgb
    return str(value)


def _serialize_font(font: Font | None) -> dict[str, Any] | None:
    if font is None:
        return None
    return {
        "name": font.name,
        "size": font.size,
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "color": _serialize_color(font.color),
    }


def _serialize_fill(fill: PatternFill | None) -> dict[str, Any] | None:
    if fill is None:
        return None
    return {
        "fill_type": fill.fill_type,
        "start_color": _serialize_color(fill.start_color),
        "end_color": _serialize_color(fill.end_color),
    }


def _serialize_alignment(alignment: Alignment) -> dict[str, Any]:
    return {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text) if alignment.wrap_text else False,
        "indent": alignment.indent or 0,
    }


def _serialize_side(side: Side) -> dict[str, Any]:
    return {
        "border_style": side.border_style,
        "color": _serialize_color(side.color),
    }


def _serialize_border(border: Border) -> dict[str, Any]:
    return {
        "left": _serialize_side(border.left),
        "right": _serialize_side(border.right),
        "top": _serialize_side(border.top),
        "bottom": _serialize_side(border.bottom),
    }


def serialize_style(spec: StyleDefinition) -> dict[str, Any]:
    """Serialize a single :class:`StyleDefinition` to a JSON-safe dict.

    Stable across openpyxl 3.1.x patch releases for the attributes we
    set explicitly — the snapshot fixture relies on this property.
    """
    return {
        "name": spec.name,
        "role": spec.role,
        "font": _serialize_font(spec.font),
        "fill": _serialize_fill(spec.fill),
        "alignment": _serialize_alignment(spec.alignment),
        "border": _serialize_border(spec.border),
        "number_format": spec.number_format,
    }


def styles_snapshot() -> dict[str, dict[str, Any]]:
    """Return ``{class_name: serialized_spec}`` for every class in :data:`STYLES`."""
    return {name: serialize_style(spec) for name, spec in STYLES.items()}


__all__ = [
    "OVERLAY_CLASSES",
    "PRIMARY_CLASSES",
    "STYLES",
    "StyleDefinition",
    "register_named_styles",
    "serialize_style",
    "styles_snapshot",
]
