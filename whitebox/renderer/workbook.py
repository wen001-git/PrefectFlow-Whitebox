"""Workbook-level entry points — render and persist.

Both functions are pure with respect to the filesystem; only
:func:`write_workbook` performs IO (a single ``Workbook.save``).
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook

from whitebox.renderer.sheet_writer import write_sheet
from whitebox.renderer.styles import register_named_styles
from whitebox.sheets.base import SheetModel


def render_workbook(sheet_models: Sequence[SheetModel]) -> Workbook:
    """Build a fully-styled ``openpyxl.Workbook`` from sheet models.

    A new workbook is created; the default ``"Sheet"`` worksheet is
    removed once at least one model has been rendered (so the result
    contains *only* the requested tabs, in the provided order). If
    ``sheet_models`` is empty the workbook is returned unchanged
    (keeps the openpyxl-default empty ``"Sheet"`` tab).
    """
    wb = Workbook()
    register_named_styles(wb)

    if not sheet_models:
        return wb

    # Drop the auto-created default sheet so output is deterministic and
    # matches the legacy renderer (which never carries an empty "Sheet").
    default_ws = wb.active
    default_title = default_ws.title if default_ws is not None else None

    for model in sheet_models:
        ws = wb.create_sheet(title=model.sheet_name)
        write_sheet(ws, model)

    if default_title is not None and default_title in wb.sheetnames:
        del wb[default_title]

    return wb


def write_workbook(wb: Workbook, path: str | Path) -> Path:
    """Persist ``wb`` to ``path`` and return the absolute :class:`Path`."""
    out = Path(path).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


__all__ = ["render_workbook", "write_workbook"]
