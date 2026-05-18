"""
STUB — replace when real engine lands.

tools/run_newsystem_mrc.py
==========================
New-system MRC validation-report runner **stub** (Round 2 C3).

This module generates a SYNTHETIC ``validation_report.xlsx`` so the
comparison harness pipeline (C4 orchestrator, C5 end-to-end harness) can
be exercised before the real whitebox engine exists.

Cross-ref: plan.md § 9 — "legacy vs new XLSX comparison" strategy pivot.

---------------------------------------------------------------------
Injected perturbations (``--mode perturbed``)
---------------------------------------------------------------------
1. **value_diff**   — Sheet ``MRC_General_Check``, row 2, column
   ``rate``  value is changed from ``0.04`` to ``0.05`` (a deliberate
   numeric discrepancy).
2. **font_diff**    — Sheet ``MRC_Advance_Check``, cell A2 font is set
   to bold=True (pristine uses bold=False).
3. **missing_row**  — Sheet ``MRC_ServiceFee_Check`` has one fewer data
   row (3 rows instead of 4) — the last data row is omitted.
4. **extra_sheet**  — An additional worksheet ``_PERTURBATION_EXTRA`` is
   appended to the workbook (not present in pristine output).

These four known-diff types map one-to-one to the four diff categories
the C5 harness is designed to detect: value, style/font, missing-row,
extra-sheet.

---------------------------------------------------------------------
Deletion policy
---------------------------------------------------------------------
Delete this stub and replace with a real engine call once
``whitebox/engine/`` (Stage 2 implementation tier) is in place and
``stage2-mrc-cell-identity-harness`` passes.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import random
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

__version__ = "1.0.0"

# ---------------------------------------------------------------------------
# Canonical MRC sheet names (from docs/mrc/1.3-sheets.zh.md)
# ---------------------------------------------------------------------------
MRC_SHEETS: list[str] = [
    "MRC_Summary_check",
    "MRC_General_Check",
    "MRC_Advance_Check",
    "MRC_ServiceFee_Check",
    "MRC_Adv_Info",
]

# Header fill for all sheets: #BCCDE9 (matches legacy renderer)
_HEADER_FILL = PatternFill("solid", fgColor="BCCDE9")
_HEADER_FONT = Font(name="Arial", size=12, bold=True)
_BODY_FONT   = Font(name="Arial", size=12, bold=False)

# ---------------------------------------------------------------------------
# Per-sheet column definitions
# ---------------------------------------------------------------------------
_SHEET_COLS: dict[str, list[str]] = {
    "MRC_Summary_check": [
        "asofdate", "total_upb", "total_scheduled_principal",
        "total_scheduled_interest", "total_actual_principal",
        "total_actual_interest", "total_deferred_principal",
        "total_deferred_interest", "total_prepayment",
        "total_net_loss", "total_servicefee",
        "total_advance_balance", "total_advance_repayment",
        "total_advance_draws",
    ],
    "MRC_General_Check": [
        "loan_id", "asofdate", "remit_type", "rate", "maturity_date",
        "upb_remit", "upb_daily", "upb_diff",
        "sched_p_remit", "sched_p_daily", "sched_p_diff",
        "sched_i_remit", "sched_i_daily", "sched_i_diff",
        "actual_p_remit", "actual_p_daily", "actual_p_diff",
        "actual_i_remit", "actual_i_daily", "actual_i_diff",
        "deferred_p_remit", "deferred_p_daily", "deferred_p_diff",
        "deferred_i_remit", "deferred_i_daily", "deferred_i_diff",
        "prepayment_remit", "prepayment_daily", "prepayment_diff",
        "net_loss_remit", "net_loss_daily", "net_loss_diff",
        "servicefee_remit", "servicefee_daily", "servicefee_diff",
    ],
    "MRC_Advance_Check": [
        "loan_id", "asofdate", "adv_bucket",
        "advance_balance_remit", "advance_balance_daily", "advance_balance_diff",
        "advance_repayment_remit", "advance_repayment_daily", "advance_repayment_diff",
        "advance_draws_remit", "advance_draws_daily", "advance_draws_diff",
        "advance_net_remit", "advance_net_daily", "advance_net_diff",
        "advance_total_remit", "advance_total_daily", "advance_total_diff",
        "advance_ytd_remit", "advance_ytd_daily", "advance_ytd_diff",
        "advance_cum_remit", "advance_cum_daily", "advance_cum_diff",
        "interest_remit", "interest_daily", "interest_diff",
    ],
    "MRC_ServiceFee_Check": [
        "loan_id", "asofdate", "upb",
        "servicefee_remit", "servicefee_portmonth", "servicefee_diff",
        "servicefee_rate", "notes",
    ],
    "MRC_Adv_Info": [
        "bucket", "description", "transaction_code",
        "amt", "amt_1m", "amt_MoM", "asofdate",
    ],
}

# ---------------------------------------------------------------------------
# Deterministic synthetic data factories
# ---------------------------------------------------------------------------

def _make_row_summary(idx: int, remit_date: str) -> list[object]:
    v = 1_000_000.0 + idx * 7_777.77
    return [
        remit_date, round(v, 2), round(v * 0.001, 2), round(v * 0.003, 2),
        round(v * 0.0008, 2), round(v * 0.0025, 2), round(v * 0.0001, 2),
        round(v * 0.0002, 2), round(v * 0.0005, 2), 0.0, round(v * 0.00025, 2),
        round(v * 0.012, 2), round(v * 0.002, 2), round(v * 0.003, 2),
    ]


def _make_row_general(idx: int, remit_date: str) -> list[object]:
    base = 50_000.0 + idx * 1_234.56
    return [
        f"LOAN{1000 + idx:04d}", remit_date, "FULL",
        0.04, "2045-01-01",
        round(base, 2), round(base, 2), 0.0,
        round(base * 0.001, 2), round(base * 0.001, 2), 0.0,
        round(base * 0.003, 2), round(base * 0.003, 2), 0.0,
        round(base * 0.0008, 2), round(base * 0.0008, 2), 0.0,
        round(base * 0.0025, 2), round(base * 0.0025, 2), 0.0,
        round(base * 0.0001, 2), round(base * 0.0001, 2), 0.0,
        round(base * 0.0002, 2), round(base * 0.0002, 2), 0.0,
        round(base * 0.0005, 2), round(base * 0.0005, 2), 0.0,
        0.0, 0.0, 0.0,
        round(base * 0.00025, 2), round(base * 0.00025, 2), 0.0,
    ]


def _make_row_advance(idx: int, remit_date: str) -> list[object]:
    base = 2_000.0 + idx * 111.11
    return [
        f"LOAN{1000 + idx:04d}", remit_date, f"BKT{idx % 3:02d}",
        round(base, 2), round(base, 2), 0.0,
        round(base * 0.1, 2), round(base * 0.1, 2), 0.0,
        round(base * 0.05, 2), round(base * 0.05, 2), 0.0,
        round(base * 0.15, 2), round(base * 0.15, 2), 0.0,
        round(base * 1.1, 2), round(base * 1.1, 2), 0.0,
        round(base * 2.0, 2), round(base * 2.0, 2), 0.0,
        round(base * 3.0, 2), round(base * 3.0, 2), 0.0,
        round(base * 0.01, 2), round(base * 0.01, 2), 0.0,
    ]


def _make_row_servicefee(idx: int, remit_date: str) -> list[object]:
    base = 50_000.0 + idx * 1_234.56
    fee = round(base * 0.00025, 2)
    return [
        f"LOAN{1000 + idx:04d}", remit_date, round(base, 2),
        fee, fee, 0.0, 0.00025, "",
    ]


def _make_row_adv_info(idx: int, remit_date: str) -> list[object]:
    amt = round(500.0 + idx * 33.33, 2)
    return [
        f"BKT{idx % 3:02d}", f"Advance description {idx}",
        f"TRX{idx:03d}", amt, round(amt * 0.95, 2),
        round(amt / max(amt * 0.95, 1) - 1, 4), remit_date,
    ]


_DATA_ROW_FACTORIES = {
    "MRC_Summary_check":    (_make_row_summary,    1),
    "MRC_General_Check":    (_make_row_general,    5),
    "MRC_Advance_Check":    (_make_row_advance,    5),
    "MRC_ServiceFee_Check": (_make_row_servicefee, 4),
    "MRC_Adv_Info":         (_make_row_adv_info,   6),
}

# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def _write_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    cols: list[str],
    rows: list[list[object]],
) -> None:
    """Write header + data rows with basic styling."""
    # Header row
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _BODY_FONT


def _build_workbook(
    remit_date: str,
    mode: str,
    seed: int,
) -> tuple[openpyxl.Workbook, list[dict[str, object]], list[dict[str, object]]]:
    """
    Build and return (workbook, sheet_stats, perturbations).
    ``perturbations`` is non-empty only in perturbed mode.
    """
    random.seed(seed)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    sheet_stats: list[dict[str, object]] = []
    perturbations: list[dict[str, object]] = []

    for sheet_name in MRC_SHEETS:
        factory_fn, n_rows = _DATA_ROW_FACTORIES[sheet_name]
        cols = _SHEET_COLS[sheet_name]

        if mode == "empty":
            rows: list[list[object]] = []
        else:
            rows = [factory_fn(i, remit_date) for i in range(n_rows)]

        # --- Perturbation 3: missing_row (ServiceFee gets one fewer row) ---
        if mode == "perturbed" and sheet_name == "MRC_ServiceFee_Check" and rows:
            rows = rows[:-1]

        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, cols, rows)

        # --- Perturbation 1: value_diff (General_Check row 2 rate → 0.05) ---
        if mode == "perturbed" and sheet_name == "MRC_General_Check" and len(rows) >= 1:
            rate_col = cols.index("rate") + 1  # 1-based
            ws.cell(row=2, column=rate_col).value = 0.05
            perturbations.append({
                "id": "value_diff",
                "sheet": "MRC_General_Check",
                "cell": f"row=2,col={rate_col}",
                "expected": 0.04,
                "actual": 0.05,
                "description": "rate changed from 0.04 to 0.05 in row 2",
            })

        # --- Perturbation 2: font_diff (Advance_Check A2 bold=True) ---
        if mode == "perturbed" and sheet_name == "MRC_Advance_Check" and len(rows) >= 1:
            cell = ws.cell(row=2, column=1)
            cell.font = Font(name="Arial", size=12, bold=True)
            perturbations.append({
                "id": "font_diff",
                "sheet": "MRC_Advance_Check",
                "cell": "A2",
                "expected": "bold=False",
                "actual": "bold=True",
                "description": "A2 font bold flipped to True",
            })

        n_data = ws.max_row - 1  # exclude header
        sheet_stats.append({
            "name": sheet_name,
            "rows": max(n_data, 0),
            "cols": len(cols),
        })

    # --- Perturbation 4: extra_sheet ---
    if mode == "perturbed":
        extra_ws = wb.create_sheet(title="_PERTURBATION_EXTRA")
        extra_ws.cell(row=1, column=1, value="extra_key")
        extra_ws.cell(row=1, column=2, value="extra_value")
        extra_ws.cell(row=2, column=1, value="injected")
        extra_ws.cell(row=2, column=2, value=True)
        perturbations.append({
            "id": "extra_sheet",
            "sheet": "_PERTURBATION_EXTRA",
            "description": "extra worksheet appended — not present in pristine output",
        })

        # Record missing_row after the ServiceFee sheet was created
        perturbations.append({
            "id": "missing_row",
            "sheet": "MRC_ServiceFee_Check",
            "description": "last data row (index 4) omitted — 3 rows instead of 4",
        })

    return wb, sheet_stats, perturbations


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="run_newsystem_mrc.py",
        description=(
            "STUB — generate a synthetic MRC validation report XLSX "
            "for comparison-harness bring-up. Replace when real engine lands."
        ),
    )
    p.add_argument("--servicer", required=True, help="Servicer code (must be 'mrc').")
    p.add_argument("--remit-date", required=True, help="Remit date (YYYY-MM-DD).")
    p.add_argument("--out-dir", required=True, help="Output directory path.")
    p.add_argument(
        "--mode",
        choices=["pristine", "perturbed", "empty"],
        default="pristine",
        help=(
            "pristine: deterministic, intended-identical to a hypothetical legacy run. "
            "perturbed: pristine + 4 known injected diffs. "
            "empty: minimal valid XLSX (no data rows)."
        ),
    )
    p.add_argument("--seed", type=int, default=42, help="RNG seed for determinism.")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:  # noqa: C901
    """Entry point. Returns exit code."""
    try:
        args = _parse_args(argv)
    except SystemExit as exc:
        return int(exc.code) if exc.code is not None else 2

    if args.servicer.lower() != "mrc":
        print(
            f"ERROR: only servicer='mrc' is supported by this stub; got {args.servicer!r}",
            file=sys.stderr,
        )
        return 2

    started_at = datetime.now(timezone.utc).isoformat()

    out_dir = Path(args.out_dir)
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        print(f"ERROR: cannot create output dir {out_dir}: {exc}", file=sys.stderr)
        return 1

    try:
        wb, sheet_stats, perturbations = _build_workbook(
            remit_date=args.remit_date,
            mode=args.mode,
            seed=args.seed,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: workbook generation failed: {exc}", file=sys.stderr)
        return 1

    xlsx_path = out_dir / "validation_report.xlsx"
    try:
        wb.save(str(xlsx_path))
    except OSError as exc:
        print(f"ERROR: cannot save XLSX: {exc}", file=sys.stderr)
        return 1

    finished_at = datetime.now(timezone.utc).isoformat()

    sha = _sha256(xlsx_path)
    size = xlsx_path.stat().st_size

    metadata: dict[str, object] = {
        "tool_version": __version__,
        "started_at": started_at,
        "finished_at": finished_at,
        "system": "newsystem-stub",
        "mode": args.mode,
        "servicer": args.servicer,
        "remit_date": args.remit_date,
        "output": {
            "xlsx_path": str(xlsx_path),
            "sha256": sha,
            "size_bytes": size,
        },
        "sheets": sheet_stats,
    }
    if perturbations:
        metadata["perturbations"] = perturbations  # type: ignore[assignment]

    meta_path = out_dir / "run_metadata.json"
    meta_path.write_text(json.dumps(metadata, indent=2), encoding="utf-8")

    if args.mode == "perturbed" and perturbations:
        pert_path = out_dir / "perturbations.json"
        pert_path.write_text(json.dumps(perturbations, indent=2), encoding="utf-8")

    print(f"OK  {xlsx_path}  sha256={sha[:12]}…  mode={args.mode}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
