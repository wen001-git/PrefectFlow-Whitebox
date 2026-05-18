"""XLSX cell-level diff tool.

How to use
----------
CLI::

    python tools/xlsx_diff.py legacy.xlsx new.xlsx \\
        [--json out.json] [--html out.html] \\
        [--float-tolerance 0.0] \\
        [--ignore-style] [--ignore-format] [--ignore-row-heights] \\
        [--max-diffs-per-sheet 1000] \\
        [--summary-only]

Python API::

    from tools.xlsx_diff import diff_workbooks, DiffReport, CellDiff, DiffOptions
    opts = DiffOptions(float_tolerance=0.001)
    report = diff_workbooks(Path("legacy.xlsx"), Path("new.xlsx"), opts)
    print(report.major_count, report.minor_count)

Exit codes::

    0 = identical (no diffs of any kind)
    1 = diffs found (minor style/format diffs only)
    2 = major diffs (structure / value / formula / merged cells)
    3 = file/IO error

Comparison dimensions (per ch 1.6 baseline contract)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
major: sheet structure, cell values, formulas, merged cell ranges
minor: number format, font (name/size/bold/italic/color/underline),
       fill (fgColor/bgColor/patternType), border (each side style+color),
       alignment (horizontal/vertical/wrap_text/indent), freeze panes,
       column widths, row heights, sheet view (tabColor, hidden)
"""

from __future__ import annotations

import datetime as dt
import hashlib
import html as html_mod
import json
import math
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import click
import openpyxl
from openpyxl.utils import get_column_letter

TOOL_VERSION = "1.0.0"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAJOR = "major"
MINOR = "minor"

CAT_VALUE = "value"
CAT_FORMULA = "formula"
CAT_FORMAT = "format"
CAT_FONT = "font"
CAT_FILL = "fill"
CAT_BORDER = "border"
CAT_ALIGNMENT = "alignment"
CAT_MERGED = "merged_cells"
CAT_FREEZE = "freeze_panes"
CAT_COL_WIDTH = "col_width"
CAT_ROW_HEIGHT = "row_height"
CAT_SHEET_VIEW = "sheet_view"
CAT_STRUCTURE = "structure"

MAJOR_CATS = {CAT_VALUE, CAT_FORMULA, CAT_MERGED, CAT_STRUCTURE}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class DiffOptions:
    """Controls which comparisons are performed and how."""

    float_tolerance: float = 0.0
    ignore_style: bool = False
    ignore_format: bool = False
    ignore_row_heights: bool = False
    max_diffs_per_sheet: int = 1000
    summary_only: bool = False


@dataclass
class CellDiff:
    """A single detected difference between two XLSX files."""

    sheet: str
    row: int  # 1-based; 0 = sheet-level attribute
    col: str  # column letter (e.g. "A"); "" = sheet-level attribute
    category: str
    severity: str
    legacy: str
    new: str
    note: str = ""

    def to_dict(self) -> dict[str, Any]:
        d: dict[str, Any] = {
            "sheet": self.sheet,
            "row": self.row,
            "col": self.col,
            "category": self.category,
            "severity": self.severity,
            "legacy": self.legacy,
            "new": self.new,
        }
        if self.note:
            d["note"] = self.note
        return d


@dataclass
class DiffReport:
    """Full diff result between two workbooks."""

    tool_version: str
    generated_at: str
    legacy_meta: dict[str, Any]
    new_meta: dict[str, Any]
    diffs: list[CellDiff]
    structure: dict[str, Any]

    @property
    def identical(self) -> bool:
        return len(self.diffs) == 0

    @property
    def major_count(self) -> int:
        return sum(1 for d in self.diffs if d.severity == MAJOR)

    @property
    def minor_count(self) -> int:
        return sum(1 for d in self.diffs if d.severity == MINOR)

    def per_sheet_summary(self) -> list[dict[str, Any]]:
        """Return per-sheet diff counts in legacy sheet order."""
        legacy_sheets: list[str] = self.structure.get("sheets_legacy", [])
        new_sheets: list[str] = self.structure.get("sheets_new", [])
        legacy_set = set(legacy_sheets)
        new_set = set(new_sheets)

        info: dict[str, dict[str, Any]] = {}
        for s in legacy_sheets + [s for s in new_sheets if s not in legacy_set]:
            info[s] = {
                "sheet": s,
                "major": 0,
                "minor": 0,
                "missing": s in legacy_set and s not in new_set,
                "extra": s not in legacy_set and s in new_set,
            }
        for d in self.diffs:
            if d.sheet in info:
                if d.severity == MAJOR:
                    info[d.sheet]["major"] += 1
                else:
                    info[d.sheet]["minor"] += 1
        # preserve order: legacy first, then extra-in-new
        ordered = [info[s] for s in legacy_sheets if s in info]
        for s in new_sheets:
            if s not in legacy_set and s in info:
                ordered.append(info[s])
        return ordered

    def to_dict(self) -> dict[str, Any]:
        return {
            "tool_version": self.tool_version,
            "generated_at": self.generated_at,
            "legacy": self.legacy_meta,
            "new": self.new_meta,
            "summary": {
                "identical": self.identical,
                "major_diff_count": self.major_count,
                "minor_diff_count": self.minor_count,
                "per_sheet": self.per_sheet_summary(),
            },
            "diffs": [d.to_dict() for d in self.diffs],
            "structure": self.structure,
        }


# ---------------------------------------------------------------------------
# File metadata helpers
# ---------------------------------------------------------------------------


def _file_meta(path: Path) -> dict[str, Any]:
    raw = path.read_bytes()
    return {
        "path": str(path.resolve()),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "size_bytes": len(raw),
        "openpyxl_version": openpyxl.__version__,
    }


# ---------------------------------------------------------------------------
# Value comparison helpers
# ---------------------------------------------------------------------------


def _norm_cell_value(v: Any) -> Any:
    """Normalise cell value to a canonical form for comparison."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v  # keep bool distinct from int
    if isinstance(v, float):
        # NaN: keep as NaN (NaN==NaN for diff purposes)
        # inf: keep as inf
        return v
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat()
    return v


def _values_equal(a: Any, b: Any, tol: float) -> bool:
    """Return True if two normalised values are considered equal."""
    na = _norm_cell_value(a)
    nb = _norm_cell_value(b)
    # NaN == NaN for diff purposes (we consider them "same" so no spurious diff)
    if isinstance(na, float) and isinstance(nb, float):
        if math.isnan(na) and math.isnan(nb):
            return True
        if math.isinf(na) and math.isinf(nb):
            return na == nb  # +inf == +inf, -inf == -inf
    if na == nb:
        return True
    # Float tolerance
    if (
        tol > 0.0
        and not isinstance(na, bool)
        and not isinstance(nb, bool)
        and isinstance(na, (int, float))
        and isinstance(nb, (int, float))
    ):
        return abs(float(na) - float(nb)) <= tol
    return False


def _value_repr(v: Any) -> str:
    """Human-readable representation for diff output."""
    if v is None:
        return "∅(None)"
    if isinstance(v, float):
        if math.isnan(v):
            return "NaN"
        if math.isinf(v):
            return "+∞" if v > 0 else "-∞"
    return repr(v)


# ---------------------------------------------------------------------------
# Formula helpers
# ---------------------------------------------------------------------------


def _norm_formula(s: Any) -> str:
    """Normalise a formula string: strip leading = and whitespace."""
    if not isinstance(s, str):
        return ""
    f = s.strip()
    if f.startswith("="):
        f = f[1:]
    return f.strip()


def _is_formula(v: Any) -> bool:
    return isinstance(v, str) and v.strip().startswith("=")


# ---------------------------------------------------------------------------
# Style attribute helpers
# ---------------------------------------------------------------------------


def _color_repr(color: Any) -> str:
    """Canonical string for an openpyxl Color object."""
    if color is None:
        return "None"
    try:
        ctype = getattr(color, "type", None)
        if ctype == "rgb" or (ctype is None and getattr(color, "rgb", None)):
            rgb = getattr(color, "rgb", None)
            if rgb and rgb != "00000000":
                return f"rgb:{rgb}"
            return "rgb:00000000"
        if ctype == "theme":
            theme = getattr(color, "theme", None)
            tint = getattr(color, "tint", 0.0)
            return f"theme:{theme},tint:{tint}"
        if ctype == "indexed":
            return f"indexed:{getattr(color, 'indexed', None)}"
        # fallback
        rgb = getattr(color, "rgb", None)
        if rgb:
            return f"rgb:{rgb}"
    except Exception:
        pass
    return "unknown"


def _font_attrs(cell: Any) -> dict[str, str]:
    f = cell.font
    if f is None:
        return {}
    return {
        "name": str(f.name) if f.name is not None else "None",
        "size": str(f.size) if f.size is not None else "None",
        "bold": str(f.bold),
        "italic": str(f.italic),
        "underline": str(f.underline) if f.underline is not None else "None",
        "color": _color_repr(f.color),
    }


def _fill_attrs(cell: Any) -> dict[str, str]:
    fl = cell.fill
    if fl is None:
        return {}
    return {
        "patternType": str(getattr(fl, "patternType", None)),
        "fgColor": _color_repr(getattr(fl, "fgColor", None)),
        "bgColor": _color_repr(getattr(fl, "bgColor", None)),
    }


def _side_repr(side: Any) -> str:
    if side is None:
        return "None"
    style = getattr(side, "border_style", None)
    color = _color_repr(getattr(side, "color", None))
    return f"style:{style},color:{color}"


def _border_attrs(cell: Any) -> dict[str, str]:
    b = cell.border
    if b is None:
        return {}
    return {
        "left": _side_repr(b.left),
        "right": _side_repr(b.right),
        "top": _side_repr(b.top),
        "bottom": _side_repr(b.bottom),
        "diagonal": _side_repr(b.diagonal),
    }


def _alignment_attrs(cell: Any) -> dict[str, str]:
    al = cell.alignment
    if al is None:
        return {}
    return {
        "horizontal": str(getattr(al, "horizontal", None)),
        "vertical": str(getattr(al, "vertical", None)),
        "wrap_text": str(getattr(al, "wrap_text", None)),
        "indent": str(getattr(al, "indent", 0)),
    }


# ---------------------------------------------------------------------------
# Per-cell diff accumulator helpers
# ---------------------------------------------------------------------------


def _diff_attrs(
    sheet: str,
    row: int,
    col: str,
    category: str,
    severity: str,
    legacy_attrs: dict[str, str],
    new_attrs: dict[str, str],
) -> list[CellDiff]:
    diffs: list[CellDiff] = []
    all_keys = set(legacy_attrs) | set(new_attrs)
    for k in sorted(all_keys):
        lv = legacy_attrs.get(k, "∅")
        nv = new_attrs.get(k, "∅")
        if lv != nv:
            diffs.append(
                CellDiff(
                    sheet=sheet,
                    row=row,
                    col=col,
                    category=category,
                    severity=severity,
                    legacy=lv,
                    new=nv,
                    note=k,
                )
            )
    return diffs


def _compare_cell_styles(
    sheet: str,
    row: int,
    col: str,
    leg_cell: Any,
    new_cell: Any,
    opts: DiffOptions,
) -> list[CellDiff]:
    """Return all style diffs for a single cell (does NOT check ignore_style here)."""
    diffs: list[CellDiff] = []
    # Number format
    if not opts.ignore_format:
        lf = str(leg_cell.number_format or "General")
        nf = str(new_cell.number_format or "General")
        if lf != nf:
            diffs.append(
                CellDiff(
                    sheet=sheet, row=row, col=col,
                    category=CAT_FORMAT, severity=MINOR,
                    legacy=lf, new=nf,
                )
            )
    if opts.ignore_style:
        return diffs
    # Font
    diffs.extend(
        _diff_attrs(sheet, row, col, CAT_FONT, MINOR, _font_attrs(leg_cell), _font_attrs(new_cell))
    )
    # Fill
    diffs.extend(
        _diff_attrs(sheet, row, col, CAT_FILL, MINOR, _fill_attrs(leg_cell), _fill_attrs(new_cell))
    )
    # Border
    diffs.extend(
        _diff_attrs(sheet, row, col, CAT_BORDER, MINOR, _border_attrs(leg_cell), _border_attrs(new_cell))
    )
    # Alignment
    diffs.extend(
        _diff_attrs(
            sheet, row, col, CAT_ALIGNMENT, MINOR,
            _alignment_attrs(leg_cell), _alignment_attrs(new_cell),
        )
    )
    return diffs


# ---------------------------------------------------------------------------
# Sheet-level diff helpers
# ---------------------------------------------------------------------------


def _compare_merged_cells(sheet: str, ws_leg: Any, ws_new: Any) -> list[CellDiff]:
    """Diff merged cell ranges as sets."""
    leg_merged = {str(r) for r in (ws_leg.merged_cells.ranges if ws_leg else [])}
    new_merged = {str(r) for r in (ws_new.merged_cells.ranges if ws_new else [])}
    diffs: list[CellDiff] = []
    for r in sorted(leg_merged - new_merged):
        diffs.append(
            CellDiff(
                sheet=sheet, row=0, col="",
                category=CAT_MERGED, severity=MAJOR,
                legacy=r, new="∅(absent)",
                note="merged range present in legacy but not in new",
            )
        )
    for r in sorted(new_merged - leg_merged):
        diffs.append(
            CellDiff(
                sheet=sheet, row=0, col="",
                category=CAT_MERGED, severity=MAJOR,
                legacy="∅(absent)", new=r,
                note="merged range present in new but not in legacy",
            )
        )
    return diffs


def _compare_freeze_panes(sheet: str, ws_leg: Any, ws_new: Any) -> list[CellDiff]:
    lf = str(ws_leg.freeze_panes) if ws_leg and ws_leg.freeze_panes else "None"
    nf = str(ws_new.freeze_panes) if ws_new and ws_new.freeze_panes else "None"
    if lf != nf:
        return [
            CellDiff(
                sheet=sheet, row=0, col="",
                category=CAT_FREEZE, severity=MINOR,
                legacy=lf, new=nf,
            )
        ]
    return []


def _compare_col_widths(sheet: str, ws_leg: Any, ws_new: Any) -> list[CellDiff]:
    diffs: list[CellDiff] = []
    leg_cols = {k: v.width for k, v in ws_leg.column_dimensions.items()} if ws_leg else {}
    new_cols = {k: v.width for k, v in ws_new.column_dimensions.items()} if ws_new else {}
    all_keys = set(leg_cols) | set(new_cols)
    for col_letter in sorted(all_keys):
        lw = str(leg_cols.get(col_letter))
        nw = str(new_cols.get(col_letter))
        if lw != nw:
            diffs.append(
                CellDiff(
                    sheet=sheet, row=0, col=col_letter,
                    category=CAT_COL_WIDTH, severity=MINOR,
                    legacy=lw, new=nw,
                )
            )
    return diffs


def _compare_row_heights(sheet: str, ws_leg: Any, ws_new: Any) -> list[CellDiff]:
    diffs: list[CellDiff] = []
    leg_rows = {k: v.height for k, v in ws_leg.row_dimensions.items()} if ws_leg else {}
    new_rows = {k: v.height for k, v in ws_new.row_dimensions.items()} if ws_new else {}
    all_keys = set(leg_rows) | set(new_rows)
    for row_num in sorted(all_keys):
        lh = str(leg_rows.get(row_num))
        nh = str(new_rows.get(row_num))
        if lh != nh:
            diffs.append(
                CellDiff(
                    sheet=sheet, row=row_num, col="",
                    category=CAT_ROW_HEIGHT, severity=MINOR,
                    legacy=lh, new=nh,
                )
            )
    return diffs


def _compare_sheet_view(sheet: str, ws_leg: Any, ws_new: Any) -> list[CellDiff]:
    diffs: list[CellDiff] = []
    # tabColor
    try:
        lc = _color_repr(ws_leg.sheet_properties.tabColor) if ws_leg else "None"
    except Exception:
        lc = "None"
    try:
        nc = _color_repr(ws_new.sheet_properties.tabColor) if ws_new else "None"
    except Exception:
        nc = "None"
    if lc != nc:
        diffs.append(
            CellDiff(
                sheet=sheet, row=0, col="",
                category=CAT_SHEET_VIEW, severity=MINOR,
                legacy=lc, new=nc, note="tabColor",
            )
        )
    # hidden state
    try:
        lh = str(getattr(ws_leg, "sheet_state", "visible")) if ws_leg else "None"
    except Exception:
        lh = "None"
    try:
        nh = str(getattr(ws_new, "sheet_state", "visible")) if ws_new else "None"
    except Exception:
        nh = "None"
    if lh != nh:
        diffs.append(
            CellDiff(
                sheet=sheet, row=0, col="",
                category=CAT_SHEET_VIEW, severity=MINOR,
                legacy=lh, new=nh, note="hidden",
            )
        )
    return diffs


# ---------------------------------------------------------------------------
# Core sheet comparison
# ---------------------------------------------------------------------------


def _compare_sheet(
    sheet: str,
    ws_leg_val: Any,
    ws_new_val: Any,
    ws_leg_fmt: Any,
    ws_new_fmt: Any,
    opts: DiffOptions,
) -> list[CellDiff]:
    """Compare two worksheet pairs (values + formats) and return sorted diffs."""
    diffs: list[CellDiff] = []
    limit = opts.max_diffs_per_sheet

    # Determine iteration range (union of both sheets' used range)
    max_row = max(
        ws_leg_fmt.max_row or 0 if ws_leg_fmt else 0,
        ws_new_fmt.max_row or 0 if ws_new_fmt else 0,
    )
    max_col = max(
        ws_leg_fmt.max_column or 0 if ws_leg_fmt else 0,
        ws_new_fmt.max_column or 0 if ws_new_fmt else 0,
    )

    for row in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)

            # Get cells (None if sheet doesn't exist)
            lvc = ws_leg_val.cell(row, col_idx) if ws_leg_val else None
            nvc = ws_new_val.cell(row, col_idx) if ws_new_val else None
            lfc = ws_leg_fmt.cell(row, col_idx) if ws_leg_fmt else None
            nfc = ws_new_fmt.cell(row, col_idx) if ws_new_fmt else None

            lval = lvc.value if lvc is not None else None
            nval = nvc.value if nvc is not None else None

            # 1. Value comparison (using data_only=True cells)
            if not _values_equal(lval, nval, opts.float_tolerance):
                diffs.append(
                    CellDiff(
                        sheet=sheet, row=row, col=col_letter,
                        category=CAT_VALUE, severity=MAJOR,
                        legacy=_value_repr(lval), new=_value_repr(nval),
                    )
                )
                if len(diffs) >= limit:
                    return _sort_diffs(diffs)

            # 2. Formula comparison (using data_only=False cells)
            if lfc is not None and nfc is not None:
                l_formula_raw = lfc.value
                n_formula_raw = nfc.value
                l_is_f = _is_formula(l_formula_raw)
                n_is_f = _is_formula(n_formula_raw)
                if l_is_f or n_is_f:
                    lf_norm = _norm_formula(l_formula_raw) if l_is_f else ""
                    nf_norm = _norm_formula(n_formula_raw) if n_is_f else ""
                    if lf_norm != nf_norm:
                        diffs.append(
                            CellDiff(
                                sheet=sheet, row=row, col=col_letter,
                                category=CAT_FORMULA, severity=MAJOR,
                                legacy=str(l_formula_raw) if l_is_f else "∅",
                                new=str(n_formula_raw) if n_is_f else "∅",
                            )
                        )
                        if len(diffs) >= limit:
                            return _sort_diffs(diffs)

            # 3. Style comparisons
            if lfc is not None and nfc is not None:
                style_diffs = _compare_cell_styles(sheet, row, col_letter, lfc, nfc, opts)
                diffs.extend(style_diffs)
                if len(diffs) >= limit:
                    return _sort_diffs(diffs)

    return _sort_diffs(diffs)


def _sort_diffs(diffs: list[CellDiff]) -> list[CellDiff]:
    """Sort diffs deterministically: row → col → category."""

    def _col_sort_key(col: str) -> int:
        if not col:
            return 0
        # Convert column letter to int for sorting
        n = 0
        for ch in col.upper():
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n

    return sorted(diffs, key=lambda d: (d.row, _col_sort_key(d.col), d.category))


# ---------------------------------------------------------------------------
# Main workbook diff function
# ---------------------------------------------------------------------------


def diff_workbooks(
    legacy_path: Path,
    new_path: Path,
    opts: DiffOptions | None = None,
) -> DiffReport:
    """Compare two XLSX workbooks and return a structured DiffReport.

    Loads each file twice: once data_only=True for computed values,
    once data_only=False for formulas and cell styles.
    """
    if opts is None:
        opts = DiffOptions()

    generated_at = dt.datetime.utcnow().isoformat() + "Z"
    legacy_meta = _file_meta(legacy_path)
    new_meta = _file_meta(new_path)

    # Load both files in both modes
    wb_leg_val = openpyxl.load_workbook(legacy_path, data_only=True, read_only=False)
    wb_new_val = openpyxl.load_workbook(new_path, data_only=True, read_only=False)
    wb_leg_fmt = openpyxl.load_workbook(legacy_path, data_only=False, read_only=False)
    wb_new_fmt = openpyxl.load_workbook(new_path, data_only=False, read_only=False)

    legacy_sheet_names: list[str] = wb_leg_fmt.sheetnames
    new_sheet_names: list[str] = wb_new_fmt.sheetnames

    all_diffs: list[CellDiff] = []

    # Structure: sheet count / order / names
    if legacy_sheet_names != new_sheet_names:
        note = (
            f"legacy=[{', '.join(legacy_sheet_names)}] "
            f"new=[{', '.join(new_sheet_names)}]"
        )
        all_diffs.append(
            CellDiff(
                sheet="(workbook)", row=0, col="",
                category=CAT_STRUCTURE, severity=MAJOR,
                legacy=str(legacy_sheet_names), new=str(new_sheet_names),
                note=note,
            )
        )

    # Union of all sheet names, legacy-order first
    legacy_set = set(legacy_sheet_names)
    new_set = set(new_sheet_names)
    all_sheets: list[str] = list(legacy_sheet_names) + [
        s for s in new_sheet_names if s not in legacy_set
    ]

    merged_diffs_per_sheet: dict[str, dict[str, list[str]]] = {}

    for sheet in all_sheets:
        in_legacy = sheet in legacy_set
        in_new = sheet in new_set

        # Missing/extra sheet
        if not in_legacy:
            all_diffs.append(
                CellDiff(
                    sheet=sheet, row=0, col="",
                    category=CAT_STRUCTURE, severity=MAJOR,
                    legacy="∅(sheet absent)", new="present",
                    note="sheet exists in new only",
                )
            )
            continue
        if not in_new:
            all_diffs.append(
                CellDiff(
                    sheet=sheet, row=0, col="",
                    category=CAT_STRUCTURE, severity=MAJOR,
                    legacy="present", new="∅(sheet absent)",
                    note="sheet exists in legacy only",
                )
            )
            continue

        ws_leg_val = wb_leg_val[sheet]
        ws_new_val = wb_new_val[sheet]
        ws_leg_fmt = wb_leg_fmt[sheet]
        ws_new_fmt = wb_new_fmt[sheet]

        # Sheet dimensions diff
        leg_dim = (ws_leg_fmt.max_row or 0, ws_leg_fmt.max_column or 0)
        new_dim = (ws_new_fmt.max_row or 0, ws_new_fmt.max_column or 0)
        if leg_dim != new_dim:
            all_diffs.append(
                CellDiff(
                    sheet=sheet, row=0, col="",
                    category=CAT_STRUCTURE, severity=MAJOR,
                    legacy=f"rows={leg_dim[0]},cols={leg_dim[1]}",
                    new=f"rows={new_dim[0]},cols={new_dim[1]}",
                    note="sheet dimensions (max_row, max_col)",
                )
            )

        # Cell-level diffs
        cell_diffs = _compare_sheet(sheet, ws_leg_val, ws_new_val, ws_leg_fmt, ws_new_fmt, opts)
        all_diffs.extend(cell_diffs)

        # Merged cells
        merged = _compare_merged_cells(sheet, ws_leg_fmt, ws_new_fmt)
        all_diffs.extend(merged)
        if merged:
            merged_diffs_per_sheet[sheet] = {
                "legacy": sorted(str(r) for r in wb_leg_fmt[sheet].merged_cells.ranges),
                "new": sorted(str(r) for r in wb_new_fmt[sheet].merged_cells.ranges),
            }

        # Freeze panes
        all_diffs.extend(_compare_freeze_panes(sheet, ws_leg_fmt, ws_new_fmt))

        # Column widths
        all_diffs.extend(_compare_col_widths(sheet, ws_leg_fmt, ws_new_fmt))

        # Row heights
        if not opts.ignore_row_heights:
            all_diffs.extend(_compare_row_heights(sheet, ws_leg_fmt, ws_new_fmt))

        # Sheet view (tabColor, hidden)
        if not opts.ignore_style:
            all_diffs.extend(_compare_sheet_view(sheet, ws_leg_fmt, ws_new_fmt))

    # Final global sort: sheet order → row → col → category
    sheet_order = {s: i for i, s in enumerate(all_sheets)}
    sheet_order["(workbook)"] = -1

    def _global_key(d: CellDiff) -> tuple[int, int, int, str]:
        si = sheet_order.get(d.sheet, len(all_sheets))
        ci = 0
        if d.col:
            n = 0
            for ch in d.col.upper():
                n = n * 26 + (ord(ch) - ord("A") + 1)
            ci = n
        return (si, d.row, ci, d.category)

    all_diffs.sort(key=_global_key)

    structure: dict[str, Any] = {
        "sheets_legacy": legacy_sheet_names,
        "sheets_new": new_sheet_names,
        "merged_diffs_per_sheet": merged_diffs_per_sheet,
    }

    return DiffReport(
        tool_version=TOOL_VERSION,
        generated_at=generated_at,
        legacy_meta=legacy_meta,
        new_meta=new_meta,
        diffs=all_diffs,
        structure=structure,
    )


# ---------------------------------------------------------------------------
# HTML report renderer
# ---------------------------------------------------------------------------

_HTML_CSS = """
body{font-family:system-ui,sans-serif;margin:2em;line-height:1.5;color:#24292e}
h1{font-size:1.5em;border-bottom:2px solid #e1e4e8;padding-bottom:.3em}
h2{font-size:1.1em;margin-top:0}
.summary-tbl{border-collapse:collapse;margin-bottom:1.5em}
.summary-tbl th,.summary-tbl td{border:1px solid #dfe2e5;padding:.3em .7em;font-size:.9em}
.summary-tbl th{background:#f6f8fa}
details{margin:.5em 0;border:1px solid #e1e4e8;border-radius:4px}
summary{padding:.5em .8em;cursor:pointer;font-weight:600;background:#f6f8fa;border-radius:4px}
summary:hover{background:#eaf0f6}
.diff-tbl{border-collapse:collapse;width:100%;margin:.5em 0}
.diff-tbl th,.diff-tbl td{border:1px solid #dfe2e5;padding:.25em .5em;font-size:.85em;text-align:left}
.diff-tbl th{background:#f6f8fa}
.major{background:#ffe0e0}
.minor{background:#fffbcc}
.badge-major{background:#d73a49;color:#fff;padding:1px 5px;border-radius:3px;font-size:.78em;font-weight:600}
.badge-minor{background:#e36209;color:#fff;padding:1px 5px;border-radius:3px;font-size:.78em;font-weight:600}
.ok{color:green;font-weight:600}
.err{color:#d73a49;font-weight:600}
code{background:#f3f3f3;padding:1px 4px;border-radius:2px;font-family:monospace;font-size:.9em}
footer{margin-top:3em;font-size:.8em;color:#6a737d;border-top:1px solid #e1e4e8;padding-top:.8em}
"""


def _he(s: str) -> str:
    return html_mod.escape(str(s))


def render_html(report: DiffReport) -> str:
    parts: list[str] = [
        "<!DOCTYPE html><html lang='en'><head>"
        "<meta charset='utf-8'>"
        f"<title>xlsx_diff — {_he(report.legacy_meta['path'])} vs {_he(report.new_meta['path'])}</title>"
        f"<style>{_HTML_CSS}</style></head><body>",
        "<h1>xlsx_diff Report</h1>",
    ]

    # File info
    parts.append(
        f"<p><strong>Legacy:</strong> <code>{_he(report.legacy_meta['path'])}</code>"
        f" (SHA-256: <code>{_he(report.legacy_meta['sha256'][:16])}…</code>)"
        f"<br><strong>New:</strong> <code>{_he(report.new_meta['path'])}</code>"
        f" (SHA-256: <code>{_he(report.new_meta['sha256'][:16])}…</code>)</p>"
    )

    # Overall status
    status_cls = "ok" if report.identical else "err"
    status_txt = "IDENTICAL" if report.identical else (
        "MAJOR DIFFS" if report.major_count else "MINOR DIFFS ONLY"
    )
    parts.append(
        f"<p>Status: <span class='{status_cls}'>{_he(status_txt)}</span> — "
        f"<strong>{report.major_count}</strong> major, "
        f"<strong>{report.minor_count}</strong> minor diffs</p>"
    )

    # Summary table
    parts.append("<h2>Per-sheet summary</h2>")
    parts.append(
        "<table class='summary-tbl'><tr><th>Sheet</th><th>Major</th><th>Minor</th>"
        "<th>Missing</th><th>Extra</th></tr>"
    )
    for ps in report.per_sheet_summary():
        row_cls = "major" if ps["major"] else ("minor" if ps["minor"] else "")
        parts.append(
            f"<tr class='{row_cls}'>"
            f"<td>{_he(ps['sheet'])}</td>"
            f"<td>{_he(str(ps['major']))}</td>"
            f"<td>{_he(str(ps['minor']))}</td>"
            f"<td>{'✓' if ps['missing'] else ''}</td>"
            f"<td>{'✓' if ps['extra'] else ''}</td>"
            "</tr>"
        )
    parts.append("</table>")

    # Per-sheet diff sections (only if not summary_only)
    if report.diffs:
        # Group diffs by sheet
        by_sheet: dict[str, list[CellDiff]] = {}
        for d in report.diffs:
            by_sheet.setdefault(d.sheet, []).append(d)

        for sheet_name, sheet_diffs in by_sheet.items():
            n_major = sum(1 for d in sheet_diffs if d.severity == MAJOR)
            n_minor = sum(1 for d in sheet_diffs if d.severity == MINOR)
            label = f"{_he(sheet_name)} — {n_major} major, {n_minor} minor"
            parts.append(f"<details><summary>{label}</summary>")
            parts.append(
                "<table class='diff-tbl'><tr>"
                "<th>Severity</th><th>Row</th><th>Col</th>"
                "<th>Category</th><th>Legacy</th><th>New</th><th>Note</th>"
                "</tr>"
            )
            for d in sheet_diffs:
                row_cls = "major" if d.severity == MAJOR else "minor"
                badge_cls = f"badge-{d.severity}"
                parts.append(
                    f"<tr class='{row_cls}'>"
                    f"<td><span class='{badge_cls}'>{_he(d.severity.upper())}</span></td>"
                    f"<td>{_he(str(d.row)) if d.row else ''}</td>"
                    f"<td>{_he(d.col)}</td>"
                    f"<td><code>{_he(d.category)}</code></td>"
                    f"<td><code>{_he(d.legacy)}</code></td>"
                    f"<td><code>{_he(d.new)}</code></td>"
                    f"<td>{_he(d.note)}</td>"
                    "</tr>"
                )
            parts.append("</table></details>")

    # Footer
    parts.append(
        f"<footer>xlsx_diff v{_he(TOOL_VERSION)} · generated {_he(report.generated_at)}"
        f" · legacy SHA-256: {_he(report.legacy_meta['sha256'])}"
        f" · new SHA-256: {_he(report.new_meta['sha256'])}"
        f" · openpyxl {_he(report.legacy_meta['openpyxl_version'])}</footer>"
    )
    parts.append("</body></html>")
    return "".join(parts)


# ---------------------------------------------------------------------------
# Exit code helper
# ---------------------------------------------------------------------------


def exit_code_for(report: DiffReport) -> int:
    """0=identical, 1=minor-only, 2=major, 3=error (never returned here)."""
    if report.identical:
        return 0
    if report.major_count:
        return 2
    return 1


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


@click.command()
@click.argument("legacy", type=click.Path(exists=True, dir_okay=False, path_type=Path))
@click.argument("new", type=click.Path(exists=True, dir_okay=False, path_type=Path))
@click.option("--json", "json_out", type=click.Path(dir_okay=False, path_type=Path), default=None,
              help="Write JSON report to this file.")
@click.option("--html", "html_out", type=click.Path(dir_okay=False, path_type=Path), default=None,
              help="Write HTML report to this file.")
@click.option("--float-tolerance", type=float, default=0.0, show_default=True,
              help="Tolerance for floating-point value comparisons.")
@click.option("--ignore-style", is_flag=True, default=False,
              help="Suppress font / fill / border / alignment diffs.")
@click.option("--ignore-format", is_flag=True, default=False,
              help="Suppress number-format diffs.")
@click.option("--ignore-row-heights", is_flag=True, default=False,
              help="Suppress row-height diffs.")
@click.option("--max-diffs-per-sheet", type=int, default=1000, show_default=True,
              help="Stop collecting cell diffs per sheet after this many.")
@click.option("--summary-only", is_flag=True, default=False,
              help="Print summary table only; omit per-cell diff details.")
def main(  # noqa: PLR0913
    legacy: Path,
    new: Path,
    json_out: Path | None,
    html_out: Path | None,
    float_tolerance: float,
    ignore_style: bool,
    ignore_format: bool,
    ignore_row_heights: bool,
    max_diffs_per_sheet: int,
    summary_only: bool,
) -> None:
    """Compare two XLSX files at cell level and report all differences.

    \b
    Exit codes:
      0  identical (no diffs)
      1  diffs found (minor style/format diffs only)
      2  major diffs (structure / value / formula / merged cells)
      3  file/IO error
    """
    opts = DiffOptions(
        float_tolerance=float_tolerance,
        ignore_style=ignore_style,
        ignore_format=ignore_format,
        ignore_row_heights=ignore_row_heights,
        max_diffs_per_sheet=max_diffs_per_sheet,
        summary_only=summary_only,
    )

    try:
        report = diff_workbooks(legacy, new, opts)
    except Exception as exc:
        click.echo(f"ERROR: {exc}", err=True)
        sys.exit(3)

    # Print summary
    click.echo(
        f"{'IDENTICAL' if report.identical else ('MAJOR' if report.major_count else 'MINOR')} "
        f"— {report.major_count} major, {report.minor_count} minor diffs"
    )
    for ps in report.per_sheet_summary():
        flags = []
        if ps["missing"]:
            flags.append("MISSING")
        if ps["extra"]:
            flags.append("EXTRA")
        flag_str = f" [{', '.join(flags)}]" if flags else ""
        click.echo(
            f"  {ps['sheet']}{flag_str}: {ps['major']} major, {ps['minor']} minor"
        )

    if not summary_only and report.diffs:
        click.echo(f"\n{len(report.diffs)} total diffs (first 20 shown):")
        for d in report.diffs[:20]:
            loc = f"{d.sheet}!{d.col}{d.row}" if d.row else f"{d.sheet}!(sheet)"
            click.echo(
                f"  [{d.severity.upper():5}] {d.category:15} @ {loc:30}  "
                f"{d.legacy!r} → {d.new!r}"
                + (f"  ({d.note})" if d.note else "")
            )
        if len(report.diffs) > 20:
            click.echo(f"  … {len(report.diffs) - 20} more (see --json / --html output)")

    # Write outputs
    if json_out:
        json_out.write_text(json.dumps(report.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
        click.echo(f"\nJSON report → {json_out}")
    if html_out:
        html_out.write_text(render_html(report), encoding="utf-8")
        click.echo(f"HTML report → {html_out}")

    sys.exit(exit_code_for(report))


if __name__ == "__main__":
    main()
